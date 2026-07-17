#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系第一轮 dry-run 的数据分析与 Excel 输出。"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pymysql.cursors import DictCursor

from lx_product_m.config import settings
from lx_product_m.services.product_write_guard import clean
from lx_product_m.sku import extract_spu

ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "reports_analysis"
DEFAULT_YEARS = [2024, 2025, 2026]
DEFAULT_STORES = ["jq_us", "rkz_us", "sy_us", "mt_us"]
FIRST_ROUND_CUTOFF = date(2024, 6, 30)
AMAZON_FOUND_HEX = "416D617A6F6E2E466F756E642E"
DETAIL_HEADERS = ["SKU", "关联 MSKU", "SPU", "开售日期", "首销店铺", "拟打值", "近半年是否有销量"]
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


def parse_years(raw: str) -> list[int]:
    try:
        years = sorted({int(x.strip()) for x in raw.split(",") if x.strip()})
    except ValueError as exc:
        raise ValueError(f"非法年份：{raw}") from exc
    if not years or any(x < 2000 or x > 2100 for x in years):
        raise ValueError(f"非法年份：{raw}")
    return years


def parse_stores(raw: str) -> list[str]:
    if raw.strip().upper() == "ALL":
        return []
    stores = sorted({x.strip().lower() for x in raw.split(",") if x.strip()})
    if not stores or any(not re.fullmatch(r"[a-z0-9_]+", x) for x in stores):
        raise ValueError(f"非法店铺：{raw}")
    return stores


def binary_hex(value: str) -> str:
    return "0x" + value.encode("utf-8").hex().upper()


def store_filter(stores: Sequence[str], alias: str = "o") -> str:
    if not stores:
        return ""
    keys = ",".join(binary_hex(x.replace("_", "-").upper()) for x in stores)
    return f"AND CAST(UPPER(REPLACE(TRIM(CAST({alias}.source AS CHAR)), '_', '-')) AS BINARY) IN ({keys})"


def listing_store_filter(stores: Sequence[str], alias: str = "l") -> str:
    if not stores:
        return ""
    keys = ",".join(binary_hex(x.replace("_", "-").upper()) for x in stores)
    return f"AND CAST(UPPER(REPLACE(TRIM(CAST({alias}.`店铺` AS CHAR)), '_', '-')) AS BINARY) IN ({keys})"


def table_exists(conn: Any, year: int) -> bool:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT COUNT(*) cnt FROM information_schema.tables "
            "WHERE table_schema='sy_order' AND table_name=%s",
            (f"total_order_{year}",),
        )
        return int(cur.fetchone()["cnt"]) > 0


def resolve_tables(conn: Any, years: Sequence[int]) -> list[str]:
    tables = []
    for year in years:
        if table_exists(conn, year):
            tables.append(f"sy_order.total_order_{year}")
        else:
            print(f"警告：sy_order.total_order_{year} 不存在，已跳过")
    if not tables:
        raise RuntimeError("未找到任何 sy_order.total_order_YYYY 表")
    return tables


def raw_order_union(tables: Sequence[str], stores: Sequence[str]) -> str:
    parts = []
    for table in tables:
        parts.append(f"""
SELECT DATE(o.`date`) order_date,
       LOWER(CONVERT(TRIM(CAST(o.source AS CHAR)) USING utf8mb4)) source,
       CAST(UPPER(REPLACE(TRIM(CAST(o.source AS CHAR)), '_', '-')) AS BINARY) store_key,
       CONVERT(TRIM(CAST(o.`order id` AS CHAR)) USING utf8mb4) order_id,
       CONVERT(TRIM(CAST(o.sku AS CHAR)) USING utf8mb4) sku,
       CAST(TRIM(CAST(o.sku AS CHAR)) AS BINARY) sku_key,
       CAST(o.quantity AS DECIMAL(18,4)) quantity,
       CAST(o.`product sales` AS DECIMAL(18,4)) product_sales
FROM {table} o
WHERE o.`date` IS NOT NULL
  AND CAST(TRIM(CAST(o.type AS CHAR)) AS BINARY) = 0x4F72646572
  AND COALESCE(CAST(o.quantity AS DECIMAL(18,4)), 0) > 0
  AND COALESCE(CAST(o.`product sales` AS DECIMAL(18,4)), 0) > 0
  AND o.sku IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(o.sku AS CHAR)))>0
  AND o.`order id` IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(o.`order id` AS CHAR)))>0
  AND LEFT(CAST(TRIM(CAST(o.`order id` AS CHAR)) AS BINARY), 3) <> 0x533031
  {store_filter(stores, 'o')}
""".strip())
    return "\nUNION ALL\n".join(parts)


def common_ctes(tables: Sequence[str], stores: Sequence[str]) -> str:
    return f"""
WITH listing_fnsku AS (
  SELECT CAST(UPPER(REPLACE(TRIM(CAST(`店铺` AS CHAR)),'_','-')) AS BINARY) store_key,
         CAST(TRIM(CAST(FNSKU AS CHAR)) AS BINARY) fnsku_key,
         MIN(CONVERT(TRIM(CAST(MSKU AS CHAR)) USING utf8mb4)) msku
  FROM lingxing.listing
  WHERE FNSKU IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(FNSKU AS CHAR)))>0
    AND MSKU IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(MSKU AS CHAR)))>0
    AND LEFT(CAST(TRIM(CAST(MSKU AS CHAR)) AS BINARY),13)<>0x{AMAZON_FOUND_HEX}
    AND SKU IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(SKU AS CHAR)))>0
  GROUP BY store_key,fnsku_key
), listing_asin AS (
  SELECT CAST(UPPER(REPLACE(TRIM(CAST(`店铺` AS CHAR)),'_','-')) AS BINARY) store_key,
         CAST(TRIM(CAST(ASIN AS CHAR)) AS BINARY) asin_key,
         MIN(CONVERT(TRIM(CAST(MSKU AS CHAR)) USING utf8mb4)) msku
  FROM lingxing.listing
  WHERE ASIN IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(ASIN AS CHAR)))>0
    AND MSKU IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(MSKU AS CHAR)))>0
    AND LEFT(CAST(TRIM(CAST(MSKU AS CHAR)) AS BINARY),13)<>0x{AMAZON_FOUND_HEX}
    AND SKU IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(SKU AS CHAR)))>0
  GROUP BY store_key,asin_key
), raw_orders AS ({raw_order_union(tables, stores)}), mapped_orders AS (
  SELECT r.order_date,r.source,r.order_id,r.sku,r.quantity,r.product_sales,
         CONVERT(CASE
           WHEN LEFT(r.sku_key, 3) = 0x583030 THEN lf.msku
           WHEN LEFT(r.sku_key, 2) = 0x4230 OR LEFT(r.sku_key, 13) = 0x{AMAZON_FOUND_HEX} THEN la.msku
           WHEN LOCATE(0x2D, r.sku_key) > 0 THEN r.sku ELSE NULL END USING utf8mb4) msku
  FROM raw_orders r
  LEFT JOIN listing_fnsku lf ON r.store_key=lf.store_key AND r.sku_key=lf.fnsku_key
  LEFT JOIN listing_asin la ON r.store_key=la.store_key
   AND (CASE WHEN LEFT(r.sku_key,13)=0x{AMAZON_FOUND_HEX} THEN SUBSTRING(r.sku_key,14) ELSE r.sku_key END)=la.asin_key
), clean_orders AS (
  SELECT order_date,source,order_id,sku,TRIM(msku) msku,
         TRIM(SUBSTRING_INDEX(msku,'-',1)) spu,quantity,product_sales
  FROM mapped_orders
  WHERE msku IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(msku AS CHAR)))>0
    AND CHAR_LENGTH(TRIM(SUBSTRING_INDEX(CAST(msku AS CHAR),'-',1)))>0
)
"""


def color_system_sql(tables: Sequence[str], stores: Sequence[str]) -> str:
    return f"""
{common_ctes(tables, stores)},
msku_first_sales AS (
  SELECT msku, MIN(order_date) AS first_sale_date FROM clean_orders GROUP BY msku
), msku_first_sale_stores AS (
  SELECT f.msku,f.first_sale_date,c.source
  FROM msku_first_sales f JOIN clean_orders c
    ON CAST(c.msku AS BINARY)=CAST(f.msku AS BINARY) AND c.order_date=f.first_sale_date
  GROUP BY f.msku,f.first_sale_date,c.source
), msku_recent_sales AS (
  SELECT msku,MAX(CASE WHEN order_date>=DATE_SUB(CURDATE(),INTERVAL 6 MONTH) THEN 1 ELSE 0 END) has_recent_sales
  FROM clean_orders GROUP BY msku
), listing_sku_msku AS (
  SELECT DISTINCT CONVERT(TRIM(CAST(l.SKU AS CHAR)) USING utf8mb4) sku,
         CAST(TRIM(CAST(l.SKU AS CHAR)) AS BINARY) sku_key,
         CONVERT(TRIM(CAST(l.MSKU AS CHAR)) USING utf8mb4) msku,
         CAST(TRIM(CAST(l.MSKU AS CHAR)) AS BINARY) msku_key
  FROM lingxing.listing l
  WHERE l.SKU IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(l.SKU AS CHAR)))>0
    AND l.MSKU IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(l.MSKU AS CHAR)))>0
    AND LEFT(CAST(TRIM(CAST(l.MSKU AS CHAR)) AS BINARY),13)<>0x{AMAZON_FOUND_HEX}
    {listing_store_filter(stores, 'l')}
), sku_opening AS (
  SELECT l.sku_key, MIN(l.sku) AS sku, MIN(f.first_sale_date) AS opening_date,
         MAX(COALESCE(r.has_recent_sales,0)) has_recent_sales
  FROM listing_sku_msku l
  LEFT JOIN msku_first_sales f ON l.msku_key=CAST(f.msku AS BINARY)
  LEFT JOIN msku_recent_sales r ON l.msku_key=CAST(r.msku AS BINARY)
  GROUP BY l.sku_key
)
SELECT o.sku,GROUP_CONCAT(DISTINCT l.msku ORDER BY l.msku SEPARATOR ',') associated_msku,
       o.opening_date,GROUP_CONCAT(DISTINCT fs.source ORDER BY fs.source SEPARATOR ',') first_sale_stores,
       o.has_recent_sales
FROM sku_opening o JOIN listing_sku_msku l ON l.sku_key=o.sku_key
LEFT JOIN msku_first_sale_stores fs ON l.msku_key=CAST(fs.msku AS BINARY) AND fs.first_sale_date=o.opening_date
GROUP BY o.sku_key,o.sku,o.opening_date,o.has_recent_sales
ORDER BY o.opening_date,o.sku
"""


def connect_analytics_db() -> Any:
    settings.validate_db()
    config = dict(settings.db_config)
    config.update(connect_timeout=60, read_timeout=7200, write_timeout=7200,
                  autocommit=True, cursorclass=DictCursor, charset="utf8mb4")
    return pymysql.connect(**config)


def run_analysis_query(conn: Any, sql: str) -> list[dict[str, Any]]:
    started = datetime.now()
    with conn.cursor() as cur:
        cur.execute("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci")
        cur.execute("SET SESSION group_concat_max_len=1048576")
        cur.execute(sql)
        rows = list(cur.fetchall())
    print(f"MSKU 首销与本地 SKU 聚合完成：{len(rows):,} 个 SKU，耗时 {(datetime.now()-started).total_seconds():.1f}s")
    return rows


def as_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def classify_opening_date(opening_date: date | None) -> str | None:
    if opening_date is None:
        return "待定"
    return "A2023" if opening_date <= FIRST_ROUND_CUTOFF else None


def prepare_analysis_rows(rows: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    prepared, skipped = [], 0
    for row in rows:
        sku = clean(row.get("sku"))
        if not sku:
            skipped += 1
            continue
        opening = as_date(row.get("opening_date"))
        proposed = classify_opening_date(opening)
        if proposed is None:
            skipped += 1
            continue
        prepared.append({
            "SKU": sku,
            "关联 MSKU": clean(row.get("associated_msku")),
            "SPU": extract_spu(sku),
            "开售日期": opening,
            "首销店铺": clean(row.get("first_sale_stores")),
            "拟打值": proposed,
            "近半年是否有销量": "是" if int(row.get("has_recent_sales") or 0) else "否",
        })
    prepared.sort(key=lambda x: (0 if x["拟打值"] == "A2023" else 1, x["开售日期"] or date.max, x["SKU"]))
    return prepared, skipped


def style_header(ws: Any, row_no: int = 1) -> None:
    for cell in ws[row_no]:
        if cell.value is not None:
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws: Any, max_width: int = 60) -> None:
    for cells in ws.columns:
        letter = get_column_letter(cells[0].column)
        width = max((len(str(c.value or "")) for c in cells[:3000]), default=8) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), max_width)


def create_dryrun_workbook(rows: Sequence[dict[str, Any]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    detail = wb.active
    detail.title = "拟打标明细"
    detail.append(DETAIL_HEADERS)
    for row in rows:
        detail.append([row.get(h) for h in DETAIL_HEADERS])
    style_header(detail)
    detail.freeze_panes, detail.auto_filter.ref = "A2", detail.dimensions
    for cell in detail["D"][1:]:
        cell.number_format = "yyyy-mm-dd"
    if detail.max_row >= 2:
        validation = DataValidation(type="list", formula1='"A2023,待定"', allow_blank=False)
        validation.error, validation.errorTitle = "拟打值只允许 A2023 或 待定", "非法拟打值"
        detail.add_data_validation(validation)
        validation.add(f"F2:F{detail.max_row}")
    auto_width(detail)

    summary = wb.create_sheet("汇总")
    summary.append(["拟打值", "SKU数"])
    counts = Counter(str(x["拟打值"]) for x in rows)
    for value in ("A2023", "待定"):
        summary.append([value, counts[value]])
    style_header(summary)
    summary.append([])
    section = summary.max_row + 1
    summary.append(["按 SPU 聚合分布"])
    summary.cell(section, 1).fill = SECTION_FILL
    summary.cell(section, 1).font = Font(name="Microsoft YaHei", bold=True)
    header = summary.max_row + 1
    summary.append(["SPU", "A2023 SKU数", "待定 SKU数", "合计"])
    by_spu: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        by_spu[str(row["SPU"])][str(row["拟打值"])] += 1
    for spu in sorted(by_spu):
        a, p = by_spu[spu]["A2023"], by_spu[spu]["待定"]
        summary.append([spu, a, p, a + p])
    style_header(summary, header)
    summary.freeze_panes = "A2"
    auto_width(summary)
    wb.save(output)


def run_dry_run(args: Any) -> int:
    years, stores = parse_years(args.years), parse_stores(args.stores)
    output = Path(args.output).expanduser().resolve() if args.output else OUTPUT_DIR / f"color_system_tagging_dryrun_{datetime.now():%Y%m%d}.xlsx"
    print("===== 颜色体系批量打标 dry-run（第一轮） =====")
    print(f"年份：{years}")
    print(f"店铺：{'ALL' if not stores else ','.join(stores)}")
    print(f"A2023 截止日期：{FIRST_ROUND_CUTOFF:%Y-%m-%d}")
    conn = connect_analytics_db()
    try:
        raw_rows = run_analysis_query(conn, color_system_sql(resolve_tables(conn, years), stores))
    finally:
        conn.close()
    targets, skipped = prepare_analysis_rows(raw_rows)
    create_dryrun_workbook(targets, output)
    counts = Counter(x["拟打值"] for x in targets)
    for row in targets[:max(0, args.show)]:
        print(f"{row['SKU']} | {row['关联 MSKU']} | {row['开售日期'] or ''} -> {row['拟打值']}")
    print(f"执行统计：处理={len(raw_rows):,} 成功=0 失败=0 跳过={skipped:,} 拟打A2023={counts['A2023']:,} 拟打待定={counts['待定']:,}")
    print(f"dry-run 清单：{output}")
    return 0
