#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系高置信度全量写入：支持 A2023 与 B2024。

安全边界：
1. prepare 只从 V3 dry-run Excel 选择高置信度、主映射精确唯一匹配记录。
2. 覆盖 P1/P2/P3，但仍排除多色套装、未收录代码、异常结构、多个品名颜色。
3. apply 写前实时读取领星：目标字段为空才写；相同值跳过；不同值拒绝覆盖。
4. 写入时完整保留品名、分类及其他自定义字段，写后再次实时复查。
5. 可重复执行；已经成功写入的 SKU 会在下一次运行时自动跳过。
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from lx_product_m.db import Database
from lx_product_m.lingxing_client import LingxingClient
from lx_product_m.services.product_write_guard import (
    PRODUCT_SET_API,
    build_guarded_product_set_body,
    clean,
    extract_custom_fields,
    fetch_live_products,
    request_with_retry,
    target_fields_match,
)
from scripts import color_system_high_confidence_round1_v2 as compat

base = compat.base

TARGET_VALUES = {"A2023", "B2024"}
DEFAULT_FIELD_ID = "207722905719915521"
SOURCE_SHEET = "全量判定明细"
REVIEW_SHEET = "拟打标明细"
REPORT_DIR = ROOT / "reports_analysis"

REQUIRED_SOURCE_HEADERS = set(base.REQUIRED_SOURCE_HEADERS)
REVIEW_HEADERS = [
    "SKU",
    "拟打值",
    "产品名称",
    "SPU",
    "颜色代码",
    "品名识别颜色",
    "处理优先级",
    "判定类型",
    "匹配方式",
    "SKU结构",
    "来源分析文件",
]


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="高置信度颜色体系全量写入（A2023+B2024）")
    sub = parser.add_subparsers(dest="command", required=True)

    prepare = sub.add_parser("prepare", help="从 V3 分析文件生成全部安全高置信度审阅清单")
    prepare.add_argument("--analysis-file", required=True)
    prepare.add_argument("--output", default="")

    apply = sub.add_parser("apply", help="写入审阅清单")
    apply.add_argument("--review-file", required=True)
    apply.add_argument("--expected-count", type=int, required=True)
    apply.add_argument("--field-id", default=DEFAULT_FIELD_ID)
    apply.add_argument("--batch-size", type=int, default=50)
    apply.add_argument("--delay", type=float, default=0.5)
    apply.add_argument("--verify-delay", type=float, default=1.0)
    apply.add_argument("--allow-outside-low-peak", action="store_true")
    return parser.parse_args(argv)


def _headers(sheet: Any) -> dict[str, int]:
    first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return {clean(value): idx for idx, value in enumerate(first) if clean(value)}


def _cell(cells: Sequence[Any], headers: dict[str, int], name: str) -> str:
    idx = headers[name]
    return clean(cells[idx] if idx < len(cells) else "")


def eligibility_reason(row: dict[str, str]) -> str:
    if row["当前颜色体系"]:
        return "当前颜色体系非空"
    if row["拟打颜色体系"] not in TARGET_VALUES:
        return "不是A2023/B2024"
    if row["置信度"] != "高":
        return "不是高置信度"
    if row["判定类型"] != "精确唯一匹配":
        return "不是主映射精确唯一匹配"
    if row["是否多色套装"] == "是":
        return "多色套装"

    priority = row["处理优先级"]
    if not any(priority.startswith(prefix) for prefix in ("P1-", "P2-", "P3-")):
        return "优先级异常"

    structure = row["SKU结构"]
    allowed_structure = (
        structure == "常规"
        or structure.startswith("颜色尺码粘连")
        or structure.startswith("特殊后缀颜色-PD")
    )
    if not allowed_structure:
        return "非安全SKU结构"

    code = row["颜色代码"]
    if not code or any(separator in code for separator in (",", "，", "、", ";", "；")):
        return "颜色代码不唯一"

    name_colors = row["品名识别颜色"]
    if not name_colors:
        return "品名未识别颜色"
    if any(separator in name_colors for separator in ("、", ",", "，", ";", "；")):
        return "品名识别多个颜色"
    return ""


def load_source_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in wb.sheetnames:
            raise ValueError(f"分析文件缺少 sheet：{SOURCE_SHEET}")
        sheet = wb[SOURCE_SHEET]
        headers = _headers(sheet)
        missing = sorted(REQUIRED_SOURCE_HEADERS - set(headers))
        if missing:
            raise ValueError("分析文件缺少列：" + ",".join(missing))

        rows: list[dict[str, str]] = []
        seen: set[str] = set()
        for cells in sheet.iter_rows(min_row=2, values_only=True):
            sku = _cell(cells, headers, "SKU")
            if not sku:
                continue
            if sku in seen:
                raise ValueError(f"分析文件存在重复SKU：{sku}")
            seen.add(sku)
            rows.append({name: _cell(cells, headers, name) for name in REQUIRED_SOURCE_HEADERS})
        return rows
    finally:
        wb.close()


def select_rows(rows: Sequence[dict[str, str]]) -> tuple[list[dict[str, str]], Counter[str]]:
    selected: list[dict[str, str]] = []
    exclusions: Counter[str] = Counter()
    for row in rows:
        reason = eligibility_reason(row)
        if reason:
            exclusions[reason] += 1
        else:
            selected.append(dict(row))

    priority_rank = {"P1": 0, "P2": 1, "P3": 2}
    system_rank = {"A2023": 0, "B2024": 1}

    def sort_key(row: dict[str, str]) -> tuple[Any, ...]:
        priority = row["处理优先级"].split("-", 1)[0]
        return (
            priority_rank.get(priority, 9),
            system_rank.get(row["拟打颜色体系"], 9),
            row["SPU"],
            row["SKU"],
        )

    selected.sort(key=sort_key)
    return selected, exclusions


def create_review_workbook(
    selected: Sequence[dict[str, str]],
    exclusions: Counter[str],
    source_path: Path,
    output: Path,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    detail = wb.active
    detail.title = REVIEW_SHEET
    detail.append(REVIEW_HEADERS)
    for row in selected:
        values = {
            **row,
            "拟打值": row["拟打颜色体系"],
            "来源分析文件": source_path.name,
        }
        detail.append([values.get(header, "") for header in REVIEW_HEADERS])
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions

    summary = wb.create_sheet("汇总")
    system_counts = Counter(row["拟打颜色体系"] for row in selected)
    priority_counts = Counter(row["处理优先级"].split("-", 1)[0] for row in selected)
    structure_counts = Counter(row["SKU结构"] for row in selected)
    summary_rows = [
        ("来源分析文件", str(source_path)),
        ("安全高置信度总数", len(selected)),
        ("A2023", system_counts["A2023"]),
        ("B2024", system_counts["B2024"]),
        ("P1", priority_counts["P1"]),
        ("P2", priority_counts["P2"]),
        ("P3", priority_counts["P3"]),
        ("普通/粘连/特殊后缀结构数", sum(structure_counts.values())),
    ]
    summary.append(["项目", "数量/内容"])
    for name, value in summary_rows:
        summary.append([name, value])
    summary.append([])
    summary.append(["排除原因", "数量"])
    for reason, count in exclusions.most_common():
        summary.append([reason, count])

    for ws in wb.worksheets:
        for column in ws.columns:
            width = min(70, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            ws.column_dimensions[column[0].column_letter].width = width
    wb.save(output)


def prepare(args: argparse.Namespace) -> int:
    analysis_path = Path(args.analysis_file).expanduser().resolve()
    if not analysis_path.exists():
        raise FileNotFoundError(f"V3分析文件不存在：{analysis_path}")
    rows = load_source_rows(analysis_path)
    selected, exclusions = select_rows(rows)
    if not selected:
        raise RuntimeError("没有符合安全条件的高置信度记录")

    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else REPORT_DIR / f"color_system_high_confidence_all_{datetime.now(base.BEIJING_TZ):%Y%m%d_%H%M%S}.xlsx"
    )
    create_review_workbook(selected, exclusions, analysis_path, output)

    systems = Counter(row["拟打颜色体系"] for row in selected)
    priorities = Counter(row["处理优先级"].split("-", 1)[0] for row in selected)
    print("===== 全部安全高置信度清单 =====")
    print(f"来源：{analysis_path}")
    print(f"总数：{len(selected):,}")
    print(f"A2023：{systems['A2023']:,}")
    print(f"B2024：{systems['B2024']:,}")
    print(f"P1/P2/P3：{priorities['P1']:,}/{priorities['P2']:,}/{priorities['P3']:,}")
    print(f"输出：{output}")
    return 0


def load_review_file(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"审阅清单不存在：{path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if REVIEW_SHEET not in wb.sheetnames:
            raise ValueError(f"审阅清单缺少 sheet：{REVIEW_SHEET}")
        sheet = wb[REVIEW_SHEET]
        headers = _headers(sheet)
        missing = [name for name in ("SKU", "拟打值") if name not in headers]
        if missing:
            raise ValueError("审阅清单缺少列：" + ",".join(missing))

        targets: list[dict[str, str]] = []
        seen: set[str] = set()
        for row_no, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            sku = _cell(cells, headers, "SKU")
            value = _cell(cells, headers, "拟打值")
            if not sku and not value:
                continue
            if not sku:
                raise ValueError(f"审阅清单第{row_no}行SKU为空")
            if value not in TARGET_VALUES:
                raise ValueError(f"审阅清单第{row_no}行拟打值非法：{value!r}")
            if sku in seen:
                raise ValueError(f"审阅清单存在重复SKU：{sku}")
            seen.add(sku)
            targets.append({"SKU": sku, "拟打值": value})
        return targets
    finally:
        wb.close()


def live_color_values(fields: Any, field_id: str) -> set[str]:
    values: set[str] = set()
    for item in fields if isinstance(fields, list) else []:
        if not isinstance(item, dict):
            continue
        item_id = clean(item.get("id"))
        name = clean(item.get("name"))
        if item_id != field_id and name != base.COLOR_FIELD_NAME:
            continue
        value = compat.field_value(item)
        if value:
            values.add(value)
    return values


async def apply_review(args: argparse.Namespace) -> int:
    if not args.allow_outside_low_peak and not base.is_beijing_low_peak():
        now = datetime.now(base.BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")
        raise RuntimeError(
            f"当前北京时间{now}不在低峰窗口00:00-06:00；"
            "请在凌晨执行，或人工确认后使用--allow-outside-low-peak"
        )

    review_path = Path(args.review_file).expanduser().resolve()
    targets = load_review_file(review_path)
    if len(targets) != int(args.expected_count):
        raise ValueError(
            f"审阅清单数量不一致：expected={int(args.expected_count):,}, actual={len(targets):,}"
        )

    counts = Counter(item["拟打值"] for item in targets)
    batch_size = max(1, min(100, int(args.batch_size)))
    field_id = base.validate_field_id(args.field_id or os.getenv(base.COLOR_FIELD_ID_ENV, ""))
    failures: list[dict[str, str]] = []
    stats: Counter[str] = Counter(processed=len(targets))

    print("===== 高置信度A2023+B2024安全写入 =====")
    print(f"审阅清单：{review_path}")
    print(f"计划处理：{len(targets):,}；A2023={counts['A2023']:,}；B2024={counts['B2024']:,}")
    print("写前保护：实时字段为空才写；相同值跳过；不同值拒绝覆盖")
    print("写入保护：完整字段合并 + product/set + 写后实时复查")

    db = Database()
    client = LingxingClient(db=db, enable_api_log=True)
    try:
        token = (await client.generate_token()).token
        if not field_id:
            field_id = base.field_id_from_snapshot(db)
        if not field_id:
            field_id, token = await base.discover_field_id_from_live(
                client, token, [item["SKU"] for item in targets]
            )
        if not field_id:
            raise RuntimeError("无法发现颜色体系字段ID")
        print(f"颜色体系字段 ID：{field_id}")

        verify_delay = max(0.0, float(args.verify_delay))
        for start in range(0, len(targets), batch_size):
            batch = targets[start:start + batch_size]
            try:
                live_map, token = await fetch_live_products(
                    client,
                    token,
                    [item["SKU"] for item in batch],
                    max_retries=base.RATE_LIMIT_RETRIES,
                    batch_size=batch_size,
                    delay_seconds=0.5,
                    retry_base_seconds=base.RATE_LIMIT_WAIT_SECONDS,
                    retry_max_seconds=base.RATE_LIMIT_WAIT_SECONDS,
                )
            except Exception as exc:
                for item in batch:
                    failures.append(base.failure_record(item["SKU"], item["拟打值"], "写前实时读取", exc))
                    stats["failed"] += 1
                base.print_apply_progress(stats, len(targets))
                continue

            pending_verify: list[tuple[str, str, dict[str, Any]]] = []
            for item in batch:
                sku, value = item["SKU"], item["拟打值"]
                target = {"id": field_id, "name": base.COLOR_FIELD_NAME, "val": value}
                try:
                    product = live_map.get(sku)
                    if not product:
                        raise RuntimeError("写前实时详情未返回该SKU")
                    fields = extract_custom_fields(product)
                    discovered = base.field_ids_from_fields(fields)
                    if discovered and discovered != {field_id}:
                        raise RuntimeError(
                            f"实时详情字段ID与配置不一致：configured={field_id}, live={sorted(discovered)}"
                        )

                    current_values = live_color_values(fields, field_id)
                    if current_values == {value}:
                        stats["skipped"] += 1
                        continue
                    if current_values:
                        message = "实时颜色体系已存在不同值，拒绝覆盖：" + ",".join(sorted(current_values))
                        failures.append(base.failure_record(sku, value, "写前冲突保护", message))
                        stats["failed"] += 1
                        continue
                    if target_fields_match(fields, [target]):
                        stats["skipped"] += 1
                        continue

                    body, _ = build_guarded_product_set_body(
                        product,
                        sku=sku,
                        target_custom_fields=[target],
                        preserve_current_category=True,
                    )
                    response, token = await request_with_retry(
                        client,
                        token,
                        PRODUCT_SET_API,
                        body,
                        max_retries=base.RATE_LIMIT_RETRIES,
                        retry_base_seconds=base.RATE_LIMIT_WAIT_SECONDS,
                        retry_max_seconds=base.RATE_LIMIT_WAIT_SECONDS,
                    )
                    if base.response_succeeded(response):
                        pending_verify.append((sku, value, body))
                    else:
                        failures.append(base.failure_record(sku, value, "product/set", response, response=response))
                        stats["failed"] += 1
                except Exception as exc:
                    failures.append(base.failure_record(sku, value, "安全合并写入", exc))
                    stats["failed"] += 1
                if args.delay > 0:
                    await asyncio.sleep(args.delay)

            if pending_verify:
                if verify_delay > 0:
                    await asyncio.sleep(verify_delay)
                try:
                    verify_map, token = await fetch_live_products(
                        client,
                        token,
                        [item[0] for item in pending_verify],
                        max_retries=base.RATE_LIMIT_RETRIES,
                        batch_size=batch_size,
                        delay_seconds=0.5,
                        retry_base_seconds=base.RATE_LIMIT_WAIT_SECONDS,
                        retry_max_seconds=base.RATE_LIMIT_WAIT_SECONDS,
                    )
                except Exception as exc:
                    for sku, value, _ in pending_verify:
                        failures.append(base.failure_record(sku, value, "写后实时复查", exc))
                        stats["failed"] += 1
                else:
                    for sku, value, body in pending_verify:
                        ok, message = base.verify_product_set_result(verify_map.get(sku), body)
                        if ok:
                            stats["success"] += 1
                        else:
                            failures.append(base.failure_record(sku, value, "写后复查", message))
                            stats["failed"] += 1
            base.print_apply_progress(stats, len(targets))
    finally:
        await client.aclose()
        db.close()

    failure_path: Path | None = None
    if failures:
        REPORT_DIR.mkdir(parents=True, exist_ok=True)
        failure_path = REPORT_DIR / f"color_system_high_confidence_all_failures_{datetime.now(base.BEIJING_TZ):%Y%m%d_%H%M%S}.xlsx"
        base.create_failure_workbook(failures, failure_path)

    print(
        f"执行统计：处理={stats['processed']:,} 成功={stats['success']:,} "
        f"失败={stats['failed']:,} 跳过={stats['skipped']:,}"
    )
    if failure_path:
        print(f"失败清单：{failure_path}")
    return 1 if failures else 0


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if args.command == "prepare":
        return prepare(args)
    return asyncio.run(apply_review(args))


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n用户中断")
        sys.exit(130)
    except Exception as exc:
        print(f"执行失败：{type(exc).__name__}: {exc}")
        raise
