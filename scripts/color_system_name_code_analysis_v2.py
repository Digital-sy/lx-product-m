#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系只读诊断 V2：修正特殊颜色后缀，并隔离未收录完整颜色代码。

本文件复用 ``color_system_name_code_analysis.py`` 的数据库查询、颜色匹配和 Excel
主体输出，只替换 SKU 解析规则，并追加“未收录颜色代码”工作表。

安全边界与原脚本一致：只读数据库、只生成 Excel、没有 --apply。
"""
from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts import color_system_name_code_analysis as base

# 已确认的特殊命名后缀。仅当“去掉后缀后的代码”确实存在于颜色映射表时才生效。
# 例：BKPD -> BK（颜色）+ PD（绑带款特殊命名）。
SPECIAL_COLOR_SUFFIXES: dict[str, str] = {
    "PD": "绑带款",
}

UNKNOWN_SHEET = "未收录颜色代码"
UNKNOWN_HEADERS = [
    "SKU",
    "产品名称",
    "SPU",
    "分类路径",
    "原始颜色段",
    "推测尺码",
    "品名识别颜色",
    "SKU结构",
    "判定原因",
    "历史是否有销量",
    "Listing记录数",
    "最早Listing创建时间",
    "处理优先级",
]

_ORIGINAL_ANALYZE_PRODUCT = base.analyze_product
_ORIGINAL_CREATE_WORKBOOK = base.create_workbook


def _normalized_parts(sku: str) -> tuple[str, list[str], list[str], str]:
    """返回 raw、全部分段、去掉款号和版型后的 body、版型。"""
    raw = base.normalize_text(sku).upper().replace("_", "-")
    raw = re.sub(r"-+", "-", raw).strip("-")
    parts = [part for part in raw.split("-") if part]
    body = parts[1:] if len(parts) >= 2 else []
    style = ""
    if body and body[0] in base.STYLE_TOKENS:
        style = body.pop(0)
    return raw, parts, body, style


def _special_suffix(token: str, index: base.MappingIndex) -> tuple[str, str, str] | None:
    """识别已确认的特殊后缀，返回（颜色代码、后缀、说明）。"""
    for suffix, description in sorted(
        SPECIAL_COLOR_SUFFIXES.items(), key=lambda item: (-len(item[0]), item[0])
    ):
        if not token.endswith(suffix) or len(token) <= len(suffix):
            continue
        code = token[: -len(suffix)]
        if code in index.codes:
            return code, suffix, description
    return None


def parse_sku(sku: str, index: base.MappingIndex) -> base.SkuParseResult:
    """V2 SKU解析。

    核心变化：
    1. ``BKPD-S`` 按已知特殊后缀解析为 ``BK + PD + S``。
    2. 只有在颜色和尺码确实处于同一个段、即 body 仅有一个段时，才允许
       “颜色代码前缀 + 尺码”拆分。
    3. ``BLP-L`` 已经存在独立尺码段 L，因此 BLP 必须作为完整颜色段校验；
       BLP 未收录时保持未知，禁止拆成 ``BL + P-L``。
    """
    raw, parts, body, style = _normalized_parts(sku)
    spu = parts[0] if parts else ""

    if not raw:
        return base.SkuParseResult("", "", "无法解析", (), "", "", False, "", "SKU为空")
    if len(parts) < 2:
        return base.SkuParseResult(
            raw, spu, "无法解析", (), "", "", False, "", "SKU缺少颜色与尺码段"
        )

    pcs_index = next(
        (
            i
            for i, part in enumerate(body)
            if re.fullmatch(r"\d+(?:PCS|PSC)", part, re.I)
        ),
        None,
    )
    if pcs_index is not None:
        pieces = body[pcs_index]
        after = body[pcs_index + 1 :]
        if len(after) < 2:
            return base.SkuParseResult(
                raw,
                spu,
                "PCS多色套装",
                (),
                "",
                style,
                True,
                pieces,
                "PCS后缺少完整颜色与尺码",
            )
        size = after[-1]
        color_parts = after[:-1]
        valid = tuple(part for part in color_parts if part in index.codes)
        unknown = [part for part in color_parts if part not in index.codes]
        if unknown:
            return base.SkuParseResult(
                raw,
                spu,
                "PCS多色套装",
                valid,
                size,
                style,
                True,
                pieces,
                "存在未知颜色段：" + ",".join(unknown),
            )
        if not valid:
            return base.SkuParseResult(
                raw,
                spu,
                "PCS多色套装",
                (),
                size,
                style,
                True,
                pieces,
                "未识别到颜色代码",
            )
        return base.SkuParseResult(
            raw, spu, "PCS多色套装", valid, size, style, True, pieces
        )

    if not body:
        return base.SkuParseResult(
            raw, spu, "无法解析", (), "", style, False, "", "缺少颜色段"
        )

    first = body[0]

    # 业务确认的特殊后缀优先于普通颜色代码判断。
    special = _special_suffix(first, index)
    if special is not None:
        code, suffix, description = special
        if len(body) < 2:
            return base.SkuParseResult(
                raw,
                spu,
                f"特殊后缀颜色-{suffix}",
                (code,),
                "",
                style,
                False,
                "",
                f"识别到特殊后缀 {suffix}（{description}），但缺少尺码段",
            )
        if len(body) > 2:
            return base.SkuParseResult(
                raw,
                spu,
                f"特殊后缀颜色-{suffix}",
                (code,),
                body[-1],
                style,
                False,
                "",
                f"识别为 {code}+{suffix}（{description}），同时存在额外分段："
                + ",".join(body[1:-1]),
            )
        return base.SkuParseResult(
            raw,
            spu,
            f"特殊后缀颜色-{suffix}（{description}）",
            (code,),
            body[1],
            style,
            False,
            "",
        )

    if first in index.codes:
        if len(body) < 2:
            return base.SkuParseResult(
                raw, spu, "常规", (first,), "", style, False, "", "缺少尺码段"
            )
        extra_codes = [part for part in body[1:-1] if part in index.codes]
        if extra_codes:
            return base.SkuParseResult(
                raw,
                spu,
                "非PCS多颜色",
                (first, *extra_codes),
                body[-1],
                style,
                True,
                "",
                "发现多个颜色代码但无PCS标记",
            )
        if len(body) > 2:
            return base.SkuParseResult(
                raw,
                spu,
                "常规位置含额外分段",
                (first,),
                body[-1],
                style,
                False,
                "",
                "尺码不会包含连字符；颜色与尺码之间存在额外分段："
                + ",".join(body[1:-1]),
            )
        return base.SkuParseResult(raw, spu, "常规", (first,), body[1], style, False, "")

    # 仅“款号-颜色代码尺码”这种没有独立尺码段的结构，允许前缀拆分。
    if len(body) == 1:
        glued = [
            code
            for code in index.codes_by_length
            if first.startswith(code) and len(first) > len(code)
        ]
        if glued:
            max_len = len(glued[0])
            best = [code for code in glued if len(code) == max_len]
            if len(best) > 1:
                return base.SkuParseResult(
                    raw,
                    spu,
                    "颜色尺码粘连",
                    tuple(best),
                    "",
                    style,
                    False,
                    "",
                    "颜色尺码粘连存在同长度代码歧义",
                )
            code = best[0]
            size = first[len(code) :]
            return base.SkuParseResult(
                raw, spu, "颜色尺码粘连", (code,), size, style, False, ""
            )

        return base.SkuParseResult(
            raw,
            spu,
            "未收录颜色代码",
            (),
            "",
            style,
            False,
            "",
            f"完整颜色段 {first} 不在内嵌映射表",
        )

    # 已有独立尺码段时，第一段必须按完整颜色代码校验，绝不按已知代码前缀强拆。
    return base.SkuParseResult(
        raw,
        spu,
        "未收录颜色代码",
        (),
        body[-1],
        style,
        False,
        "",
        f"完整颜色段 {first} 不在内嵌映射表；已存在独立尺码段 {body[-1]}，禁止按前缀拆分",
    )


def _raw_color_segment(sku: str) -> str:
    _, _, body, _ = _normalized_parts(sku)
    if not body:
        return ""
    pcs_index = next(
        (
            i
            for i, part in enumerate(body)
            if re.fullmatch(r"\d+(?:PCS|PSC)", part, re.I)
        ),
        None,
    )
    if pcs_index is not None:
        after = body[pcs_index + 1 :]
        return ",".join(after[:-1]) if len(after) >= 2 else ""
    return body[0]


def analyze_product(
    source: dict[str, Any],
    index: base.MappingIndex,
    sales_stats: dict[str, dict[str, Any]],
    listing_stats: dict[str, dict[str, Any]],
    sales_status: str,
) -> dict[str, Any]:
    row = _ORIGINAL_ANALYZE_PRODUCT(
        source, index, sales_stats, listing_stats, sales_status
    )
    parsed = parse_sku(base.clean(source.get("sku")), index)
    raw_segment = _raw_color_segment(base.clean(source.get("sku")))
    row["_原始颜色段"] = raw_segment

    if parsed.structure == "未收录颜色代码":
        row["判定类型"] = "未收录颜色代码"
        row["匹配方式"] = "完整颜色段校验"
        row["判定原因"] = parsed.error
        row["拟打颜色体系"] = "待定"
        row["置信度"] = "低"
        row["是否允许未来自动写入"] = "否"

    if parsed.structure.startswith("特殊后缀颜色-"):
        suffix = parsed.structure.split("-", 1)[1]
        row["匹配方式"] = f"特殊后缀修正（{suffix}）+{row['匹配方式']}"
        row["判定原因"] = (
            f"原始颜色段 {raw_segment} 按特殊规则解析为颜色代码 "
            f"{','.join(parsed.color_codes)}；{row['判定原因']}"
        )

    return row


def _append_unknown_sheet(
    output: Path, detail_rows: Sequence[dict[str, Any]]
) -> None:
    unknown_rows = [
        row for row in detail_rows if row.get("判定类型") == "未收录颜色代码"
    ]

    wb = load_workbook(output)
    try:
        if UNKNOWN_SHEET in wb.sheetnames:
            del wb[UNKNOWN_SHEET]
        ws = wb.create_sheet(UNKNOWN_SHEET)
        ws.append(UNKNOWN_HEADERS)

        header_fill = PatternFill("solid", fgColor="C65911")
        header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for row in unknown_rows:
            ws.append(
                [
                    row.get("SKU", ""),
                    row.get("产品名称", ""),
                    row.get("SPU", ""),
                    row.get("分类路径", ""),
                    row.get("_原始颜色段", ""),
                    row.get("尺码", ""),
                    row.get("品名识别颜色", ""),
                    row.get("SKU结构", ""),
                    row.get("判定原因", ""),
                    row.get("历史是否有销量", ""),
                    row.get("Listing记录数", 0),
                    row.get("最早Listing创建时间", ""),
                    row.get("处理优先级", ""),
                ]
            )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:M{len(unknown_rows) + 1}"
        widths = [24, 42, 14, 28, 18, 14, 24, 34, 70, 16, 14, 22, 24]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

        summary = wb["汇总"]
        summary.append(["未收录颜色代码（完整段）", len(unknown_rows)])
        wb.save(output)
    finally:
        wb.close()


def create_workbook(
    detail_rows: list[dict[str, Any]],
    backtest_rows: list[dict[str, Any]],
    index: base.MappingIndex,
    sales_status: str,
    output: Path,
) -> None:
    _ORIGINAL_CREATE_WORKBOOK(
        detail_rows, backtest_rows, index, sales_status, output
    )
    _append_unknown_sheet(output, detail_rows)


def install_overrides() -> None:
    base.parse_sku = parse_sku
    base.analyze_product = analyze_product
    base.create_workbook = create_workbook


def main(argv: Sequence[str] | None = None) -> int:
    install_overrides()
    return base.main(argv)


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n用户中断")
        sys.exit(130)
