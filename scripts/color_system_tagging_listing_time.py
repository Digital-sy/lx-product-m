#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系补标：按 SKU 的最早 Listing 创建时间继续判定 A2023。

默认只生成审阅 Excel。正式写入只提交拟打值为 A2023 的行，并复用
color_system_tagging.py 的写前实时读取、完整字段合并和写后复查保护。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from lx_product_m.db import Database
from lx_product_m.services.product_write_guard import clean
from scripts.color_system_tagging import (
    BEIJING_TZ,
    COLOR_FIELD_ID_ENV,
    COLOR_FIELD_NAME,
    OUTPUT_DIR,
    apply_review_file,
    auto_width,
    style_header,
)

DEFAULT_CUTOFF = date(2024, 6, 30)
DEFAULT_LISTING_SCHEMA = "lingxing"
DEFAULT_LISTING_TABLE = "listing"
DEFAULT_EXCLUDE_PREFIXES = ("XH", "FAB", "LCS", "PF")
DETAIL_HEADERS = [
    "SKU",
    "产品名称",
    "SPU",
    "当前颜色体系",
    "最早Listing创建时间",
    "拟打值",
    "判定原因",
    "前缀剔除",
]

CREATED_COLUMN_EXACT_PRIORITY = (
    "listing_create_time",
    "listing_created_time",
    "listing_created_at",
    "listing_create_date",
    "listingcreatetime",
    "listingcreatedtime",
    "listingcreatedat",
    "listing创建时间",
    "创建时间",
    "创建日期",
    "create_time",
    "created_time",
    "created_at",
    "create_date",
    "上架时间",
    "刊登时间",
)
DATE_DATA_TYPES = {"date", "datetime", "timestamp"}


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="颜色体系补标（按最早 Listing 创建时间判定 A2023）")
    parser.add_argument("--cutoff", default=DEFAULT_CUTOFF.isoformat(), help="A2023 截止日期，默认2024-06-30（含）")
    parser.add_argument("--listing-schema", default=DEFAULT_LISTING_SCHEMA)
    parser.add_argument("--listing-table", default=DEFAULT_LISTING_TABLE)
    parser.add_argument("--listing-created-column", default="", help="Listing 创建时间列；留空时自动识别")
    parser.add_argument(
        "--exclude-prefixes",
        default=",".join(DEFAULT_EXCLUDE_PREFIXES),
        help="从最终不确定数量中剔除的 SKU 前缀，默认XH,FAB,LCS,PF",
    )
    parser.add_argument("--output", default="", help="dry-run Excel 输出路径")
    parser.add_argument("--show", type=int, default=30)
    parser.add_argument("--apply", action="store_true", help="写入人工审阅后的清单")
    parser.add_argument("--review-file", default="", help="--apply 时必填")
    parser.add_argument("--field-id", default="", help=f"领星颜色体系字段 ID；默认读取 {COLOR_FIELD_ID_ENV}")
    parser.add_argument("--batch-size", type=int, default=100)
    parser.add_argument("--delay", type=float, default=0.5)
    parser.add_argument("--verify-delay", type=float, default=1.0)
    parser.add_argument("--allow-outside-low-peak", action="store_true")
    return parser.parse_args(argv)


def parse_cutoff(raw: str) -> date:
    try:
        return datetime.strptime(clean(raw), "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"非法截止日期：{raw!r}，必须为 YYYY-MM-DD") from exc


def parse_prefixes(raw: str) -> tuple[str, ...]:
    values: list[str] = []
    for item in str(raw or "").split(","):
        value = clean(item).upper()
        if value and value not in values:
            values.append(value)
    if not values:
        raise ValueError("--exclude-prefixes 不能为空")
    return tuple(values)


def quote_identifier(value: str) -> str:
    return "`" + str(value).replace("`", "``") + "`"


def normalize_column_name(value: str) -> str:
    return re.sub(r"[\s_\-]+", "", clean(value)).lower()


def load_table_columns(db: Database, schema: str, table: str) -> list[dict[str, str]]:
    return db.fetch_all(
        """
        SELECT COLUMN_NAME column_name, DATA_TYPE data_type, COLUMN_TYPE column_type
        FROM information_schema.columns
        WHERE table_schema=%s AND table_name=%s
        ORDER BY ORDINAL_POSITION
        """,
        (schema, table),
    )


def resolve_column(columns: Sequence[dict[str, Any]], requested: str, *, kind: str) -> str:
    names = [clean(row.get("column_name")) for row in columns if clean(row.get("column_name"))]
    by_lower = {name.lower(): name for name in names}
    if requested:
        actual = by_lower.get(clean(requested).lower())
        if not actual:
            raise RuntimeError(f"{kind}列不存在：{requested!r}；实际列={names}")
        return actual

    if kind == "SKU":
        for candidate in ("sku", "local_sku", "本地sku", "本地SKU"):
            actual = by_lower.get(candidate.lower())
            if actual:
                return actual
        raise RuntimeError(f"无法自动识别 SKU 列；实际列={names}")

    normalized = {normalize_column_name(name): name for name in names}
    for candidate in CREATED_COLUMN_EXACT_PRIORITY:
        actual = normalized.get(normalize_column_name(candidate))
        if actual:
            return actual

    scored: list[tuple[int, str]] = []
    type_by_name = {clean(row.get("column_name")): clean(row.get("data_type")).lower() for row in columns}
    for name in names:
        if type_by_name.get(name) not in DATE_DATA_TYPES:
            continue
        norm = normalize_column_name(name)
        score = 0
        if "listing" in norm:
            score += 5
        if "创建" in name or "create" in norm or "created" in norm:
            score += 4
        if "上架" in name or "刊登" in name:
            score += 3
        if "update" in norm or "更新" in name or "modify" in norm:
            score -= 10
        if score > 0:
            scored.append((score, name))
    scored.sort(reverse=True)
    if not scored:
        date_columns = [name for name in names if type_by_name.get(name) in DATE_DATA_TYPES]
        raise RuntimeError(
            "无法自动识别 Listing 创建时间列；"
            f"日期类列={date_columns}。请通过 --listing-created-column 显式指定"
        )
    top_score = scored[0][0]
    top = [name for score, name in scored if score == top_score]
    if len(top) != 1:
        raise RuntimeError(f"Listing 创建时间列存在歧义：{top}；请通过 --listing-created-column 显式指定")
    return top[0]


def parse_custom_fields(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError:
            return []
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def original_field_value(item: dict[str, Any]) -> str:
    for key in ("val_text", "val", "value", "field_value"):
        if key in item and item.get(key) is not None:
            return clean(item.get(key))
    return ""


def extract_color_system(value: Any) -> str:
    values: list[str] = []
    for item in parse_custom_fields(value):
        if clean(item.get("name")) == COLOR_FIELD_NAME:
            field_value = original_field_value(item)
            if field_value and field_value not in values:
                values.append(field_value)
    if not values:
        return ""
    if len(values) == 1:
        return values[0]
    return "[多值]" + "|".join(values)


def as_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        raw = float(value)
        if raw > 10_000_000_000:
            raw /= 1000
        try:
            return datetime.fromtimestamp(raw, tz=timezone.utc).date()
        except (OverflowError, OSError, ValueError):
            return None
    text = clean(value)
    match = re.search(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    if re.fullmatch(r"\d{8}", text):
        try:
            return datetime.strptime(text, "%Y%m%d").date()
        except ValueError:
            return None
    return None


def prefix_of(sku: str, prefixes: Sequence[str]) -> str:
    upper = clean(sku).upper()
    for prefix in prefixes:
        if upper.startswith(prefix):
            return prefix
    return ""


def load_snapshot_rows(db: Database) -> list[dict[str, Any]]:
    return db.fetch_all(
        """
        SELECT sku, product_name, spu, custom_fields_json
        FROM lxpm_product_category_snapshot
        WHERE sku IS NOT NULL AND CHAR_LENGTH(TRIM(CAST(sku AS CHAR))) > 0
        ORDER BY sku
        """
    )


def load_earliest_listing_times(
    db: Database,
    schema: str,
    table: str,
    sku_column: str,
    created_column: str,
) -> dict[str, Any]:
    sql = f"""
        SELECT CONVERT(TRIM(CAST({quote_identifier(sku_column)} AS CHAR)) USING utf8mb4) sku,
               MIN({quote_identifier(created_column)}) earliest_listing_create_time
        FROM {quote_identifier(schema)}.{quote_identifier(table)}
        WHERE {quote_identifier(sku_column)} IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST({quote_identifier(sku_column)} AS CHAR))) > 0
          AND {quote_identifier(created_column)} IS NOT NULL
        GROUP BY CAST(TRIM(CAST({quote_identifier(sku_column)} AS CHAR)) AS BINARY)
    """
    rows = db.fetch_all(sql)
    return {
        clean(row.get("sku")): row.get("earliest_listing_create_time")
        for row in rows
        if clean(row.get("sku"))
    }


def prepare_rows(
    snapshot_rows: Iterable[dict[str, Any]],
    earliest_by_sku: dict[str, Any],
    cutoff: date,
    exclude_prefixes: Sequence[str],
) -> tuple[list[dict[str, Any]], Counter[str], Counter[str]]:
    rows: list[dict[str, Any]] = []
    stats: Counter[str] = Counter()
    excluded: Counter[str] = Counter()
    for source in snapshot_rows:
        sku = clean(source.get("sku"))
        if not sku:
            continue
        stats["snapshot_total"] += 1
        current = extract_color_system(source.get("custom_fields_json"))
        if current == "A2023":
            stats["existing_a2023"] += 1
            continue
        if current == "B2024":
            stats["existing_b2024"] += 1
            continue
        if current:
            stats["existing_other"] += 1
            continue

        stats["blank_color"] += 1
        raw_time = earliest_by_sku.get(sku)
        earliest = as_date(raw_time)
        matched_prefix = prefix_of(sku, exclude_prefixes)
        if earliest is not None and earliest <= cutoff:
            proposed = "A2023"
            reason = f"最早Listing创建时间≤{cutoff.isoformat()}"
            stats["listing_a2023"] += 1
        else:
            proposed = "待定"
            if raw_time in (None, ""):
                reason = "无Listing创建时间"
                stats["uncertain_no_listing_time"] += 1
            elif earliest is None:
                reason = f"Listing创建时间无法解析：{clean(raw_time)}"
                stats["uncertain_invalid_listing_time"] += 1
            else:
                reason = f"最早Listing创建时间>{cutoff.isoformat()}"
                stats["uncertain_after_cutoff"] += 1
            stats["uncertain_total"] += 1
            if matched_prefix:
                excluded[matched_prefix] += 1
                stats["uncertain_excluded_prefix"] += 1

        rows.append(
            {
                "SKU": sku,
                "产品名称": clean(source.get("product_name")),
                "SPU": clean(source.get("spu")),
                "当前颜色体系": current,
                "最早Listing创建时间": earliest,
                "拟打值": proposed,
                "判定原因": reason,
                "前缀剔除": matched_prefix if proposed == "待定" else "",
            }
        )
    stats["uncertain_after_exclusion"] = stats["uncertain_total"] - stats["uncertain_excluded_prefix"]
    rows.sort(
        key=lambda row: (
            0 if row["拟打值"] == "A2023" else 1,
            row["前缀剔除"],
            row["最早Listing创建时间"] or date.max,
            row["SKU"],
        )
    )
    return rows, stats, excluded


def append_summary_row(sheet: Any, name: str, value: Any) -> None:
    sheet.append([name, value])


def create_workbook(
    rows: Sequence[dict[str, Any]],
    stats: Counter[str],
    excluded: Counter[str],
    output: Path,
    *,
    cutoff: date,
    listing_source: str,
    created_column: str,
    exclude_prefixes: Sequence[str],
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    detail = wb.active
    detail.title = "拟打标明细"
    detail.append(DETAIL_HEADERS)
    for row in rows:
        detail.append([row.get(header, "") for header in DETAIL_HEADERS])
    style_header(detail)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    if detail.max_row >= 2:
        validation = DataValidation(type="list", formula1='"A2023,待定"', allow_blank=False)
        validation.error = "拟打值只允许 A2023 或 待定"
        validation.errorTitle = "非法拟打值"
        detail.add_data_validation(validation)
        value_col = DETAIL_HEADERS.index("拟打值") + 1
        validation.add(f"{detail.cell(1, value_col).column_letter}2:{detail.cell(1, value_col).column_letter}{detail.max_row}")
    auto_width(detail, max_width=80)

    summary = wb.create_sheet("汇总")
    summary.append(["指标", "数量/值"])
    style_header(summary)
    append_summary_row(summary, "Listing来源", listing_source)
    append_summary_row(summary, "Listing创建时间字段", created_column)
    append_summary_row(summary, "A2023截止日期（含）", cutoff.isoformat())
    append_summary_row(summary, "产品快照SKU总数", stats["snapshot_total"])
    append_summary_row(summary, "已有A2023", stats["existing_a2023"])
    append_summary_row(summary, "已有B2024", stats["existing_b2024"])
    append_summary_row(summary, "其他非空颜色体系（不覆盖）", stats["existing_other"])
    append_summary_row(summary, "颜色体系为空", stats["blank_color"])
    append_summary_row(summary, "按Listing时间明确A2023", stats["listing_a2023"])
    append_summary_row(summary, "剩余不确定（剔除前）", stats["uncertain_total"])
    append_summary_row(summary, "其中无Listing创建时间", stats["uncertain_no_listing_time"])
    append_summary_row(summary, "其中Listing时间无法解析", stats["uncertain_invalid_listing_time"])
    append_summary_row(summary, "其中Listing时间晚于截止日期", stats["uncertain_after_cutoff"])
    for prefix in exclude_prefixes:
        append_summary_row(summary, f"不确定中{prefix}前缀", excluded[prefix])
    append_summary_row(summary, "四类前缀剔除合计", stats["uncertain_excluded_prefix"])
    append_summary_row(summary, "剩余不确定（剔除后）", stats["uncertain_after_exclusion"])
    summary.freeze_panes = "A2"
    auto_width(summary, max_width=80)

    uncertain = wb.create_sheet("剩余不确定")
    uncertain.append(DETAIL_HEADERS)
    for row in rows:
        if row["拟打值"] == "待定" and not row["前缀剔除"]:
            uncertain.append([row.get(header, "") for header in DETAIL_HEADERS])
    style_header(uncertain)
    uncertain.freeze_panes = "A2"
    uncertain.auto_filter.ref = uncertain.dimensions
    auto_width(uncertain, max_width=80)

    excluded_sheet = wb.create_sheet("前缀剔除明细")
    excluded_sheet.append(DETAIL_HEADERS)
    for row in rows:
        if row["拟打值"] == "待定" and row["前缀剔除"]:
            excluded_sheet.append([row.get(header, "") for header in DETAIL_HEADERS])
    style_header(excluded_sheet)
    excluded_sheet.freeze_panes = "A2"
    excluded_sheet.auto_filter.ref = excluded_sheet.dimensions
    auto_width(excluded_sheet, max_width=80)

    wb.save(output)


def run_dry_run(args: argparse.Namespace) -> int:
    cutoff = parse_cutoff(args.cutoff)
    prefixes = parse_prefixes(args.exclude_prefixes)
    db = Database()
    try:
        columns = load_table_columns(db, args.listing_schema, args.listing_table)
        if not columns:
            raise RuntimeError(f"Listing表不存在或无字段：{args.listing_schema}.{args.listing_table}")
        sku_column = resolve_column(columns, "", kind="SKU")
        created_column = resolve_column(columns, args.listing_created_column, kind="Listing创建时间")
        print(f"Listing表：{args.listing_schema}.{args.listing_table}")
        print(f"SKU字段：{sku_column}")
        print(f"Listing创建时间字段：{created_column}")
        snapshot_rows = load_snapshot_rows(db)
        earliest_by_sku = load_earliest_listing_times(
            db, args.listing_schema, args.listing_table, sku_column, created_column
        )
    finally:
        db.close()

    rows, stats, excluded = prepare_rows(snapshot_rows, earliest_by_sku, cutoff, prefixes)
    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else OUTPUT_DIR / f"color_system_tagging_listing_time_dryrun_{datetime.now(BEIJING_TZ):%Y%m%d_%H%M%S}.xlsx"
    )
    create_workbook(
        rows,
        stats,
        excluded,
        output,
        cutoff=cutoff,
        listing_source=f"{args.listing_schema}.{args.listing_table}",
        created_column=created_column,
        exclude_prefixes=prefixes,
    )
    print("===== Listing 创建时间补标 dry-run =====")
    print(f"产品快照 SKU：{stats['snapshot_total']:,}")
    print(f"颜色体系为空：{stats['blank_color']:,}")
    print(f"按 Listing 时间明确 A2023：{stats['listing_a2023']:,}")
    print(f"剩余不确定（剔除前）：{stats['uncertain_total']:,}")
    for prefix in prefixes:
        print(f"  剔除 {prefix} 前缀：{excluded[prefix]:,}")
    print(f"四类前缀剔除合计：{stats['uncertain_excluded_prefix']:,}")
    print(f"剩余不确定（剔除后）：{stats['uncertain_after_exclusion']:,}")
    print(f"输出：{output}")
    show = max(0, int(args.show))
    if show:
        print("\n前几条拟打 A2023：")
        for row in [x for x in rows if x["拟打值"] == "A2023"][:show]:
            print(f"{row['SKU']}\t{row['最早Listing创建时间']}\tA2023")
    return 0


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if args.apply:
        return asyncio.run(
            apply_review_file(
                args,
                write_values={"A2023"},
                failure_prefix="color_system_tagging_listing_time_failures",
            )
        )
    if args.review_file:
        raise ValueError("--review-file 只允许与 --apply 一起使用")
    return run_dry_run(args)


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n用户中断")
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001
        print(f"执行失败：{type(exc).__name__}: {exc}")
        raise
