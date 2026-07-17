#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系批量打标（第一轮：A2023 存量判定）。

默认只生成审阅清单；只有显式指定 ``--apply --review-file`` 才会写入领星。
写入复用 product_write_guard 的实时读取、完整字段合并和 product/set 保护链路。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

import pymysql
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pymysql.cursors import DictCursor

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from lx_product_m.config import settings
from lx_product_m.db import Database
from lx_product_m.lingxing_client import LingxingClient
from lx_product_m.services.product_write_guard import (
    PRODUCT_SET_API,
    build_guarded_product_set_body,
    clean,
    extract_custom_fields,
    fetch_live_products,
    request_with_retry,
    target_fields_match,
)
from lx_product_m.sku import extract_spu

OUTPUT_DIR = ROOT / "reports_analysis"
DEFAULT_YEARS = [2024, 2025, 2026]
DEFAULT_STORES = ["jq_us", "rkz_us", "sy_us", "mt_us"]
COLOR_FIELD_NAME = "颜色体系"
COLOR_FIELD_ID_ENV = "LX_COLOR_SYSTEM_FIELD_ID"
TARGET_VALUES = {"A2023", "待定"}
FIRST_ROUND_CUTOFF = date(2024, 6, 30)
AMAZON_FOUND_HEX = "416D617A6F6E2E466F756E642E"  # Amazon.Found.
BEIJING_TZ = timezone(timedelta(hours=8), name="Asia/Shanghai")
LOW_PEAK_START = time(0, 0)
LOW_PEAK_END = time(6, 0)
RATE_LIMIT_RETRIES = 3
RATE_LIMIT_WAIT_SECONDS = 30.0

DETAIL_HEADERS = [
    "SKU",
    "关联 MSKU",
    "SPU",
    "开售日期",
    "首销店铺",
    "拟打值",
    "近半年是否有销量",
]
FAILURE_HEADERS = ["SKU", "拟打值", "失败阶段", "错误码", "错误信息", "request_id"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="颜色体系批量打标（第一轮：A2023 存量判定）")
    parser.add_argument("--years", default=",".join(map(str, DEFAULT_YEARS)))
    parser.add_argument(
        "--stores",
        default=",".join(DEFAULT_STORES),
        help="店铺 source，逗号分隔；ALL 表示全部店铺",
    )
    parser.add_argument("--output", default="", help="dry-run Excel 输出路径")
    parser.add_argument("--apply", action="store_true", help="写入人工审阅后的清单")
    parser.add_argument("--review-file", default="", help="--apply 时必填：审阅后的 dry-run Excel")
    parser.add_argument(
        "--field-id",
        default="",
        help=f"颜色体系字段 ID；默认尝试从 {COLOR_FIELD_ID_ENV}、快照或实时详情发现",
    )
    parser.add_argument("--batch-size", type=int, default=100, help="写前实时读取批次，最大100")
    parser.add_argument("--delay", type=float, default=0.5, help="每个 SKU 写入后的等待秒数")
    parser.add_argument("--show", type=int, default=30, help="dry-run 控制台预览行数")
    parser.add_argument(
        "--allow-outside-low-peak",
        action="store_true",
        help="允许在北京时间 00:00-06:00 之外执行 --apply",
    )
    return parser.parse_args(argv)


def parse_years(raw: str) -> list[int]:
    try:
        years = sorted({int(item.strip()) for item in raw.split(",") if item.strip()})
    except ValueError as exc:
        raise ValueError(f"非法年份：{raw}") from exc
    if not years or any(year < 2000 or year > 2100 for year in years):
        raise ValueError(f"非法年份：{raw}")
    return years


def parse_stores(raw: str) -> list[str]:
    if raw.strip().upper() == "ALL":
        return []
    stores = sorted({item.strip().lower() for item in raw.split(",") if item.strip()})
    if not stores or any(not re.fullmatch(r"[a-z0-9_]+", item) for item in stores):
        raise ValueError(f"非法店铺：{raw}")
    return stores


def binary_hex(value: str) -> str:
    return "0x" + value.encode("utf-8").hex().upper()


def store_filter(stores: Sequence[str], alias: str = "o") -> str:
    if not stores:
        return ""
    keys = ",".join(binary_hex(item.replace("_", "-").upper()) for item in stores)
    return f"""
      AND CAST(UPPER(REPLACE(TRIM(CAST({alias}.source AS CHAR)), '_', '-')) AS BINARY)
          IN ({keys})"""


def listing_store_filter(stores: Sequence[str], alias: str = "l") -> str:
    if not stores:
        return ""
    keys = ",".join(binary_hex(item.replace("_", "-").upper()) for item in stores)
    return f"""
      AND CAST(UPPER(REPLACE(TRIM(CAST({alias}.`店铺` AS CHAR)), '_', '-')) AS BINARY)
          IN ({keys})"""


def table_exists(conn: Any, year: int) -> bool:
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT COUNT(*) cnt FROM information_schema.tables "
            "WHERE table_schema='sy_order' AND table_name=%s",
            (f"total_order_{year}",),
        )
        return int(cursor.fetchone()["cnt"]) > 0


def resolve_tables(conn: Any, years: Sequence[int]) -> list[str]:
    tables: list[str] = []
    for year in years:
        if table_exists(conn, year):
            tables.append(f"sy_order.total_order_{year}")
        else:
            print(f"警告：sy_order.total_order_{year} 不存在，已跳过")
    if not tables:
        raise RuntimeError("未找到任何 sy_order.total_order_YYYY 表")
    return tables


def raw_order_union(tables: Sequence[str], stores: Sequence[str]) -> str:
    parts: list[str] = []
    for table in tables:
        parts.append(
            f"""
        SELECT
            DATE(o.`date`) AS order_date,
            LOWER(CONVERT(TRIM(CAST(o.source AS CHAR)) USING utf8mb4)) AS source,
            CAST(UPPER(REPLACE(TRIM(CAST(o.source AS CHAR)), '_', '-')) AS BINARY) AS store_key,
            CONVERT(TRIM(CAST(o.`order id` AS CHAR)) USING utf8mb4) AS order_id,
            CONVERT(TRIM(CAST(o.sku AS CHAR)) USING utf8mb4) AS sku,
            CAST(TRIM(CAST(o.sku AS CHAR)) AS BINARY) AS sku_key,
            CAST(o.quantity AS DECIMAL(18,4)) AS quantity,
            CAST(o.`product sales` AS DECIMAL(18,4)) AS product_sales
        FROM {table} o
        WHERE o.`date` IS NOT NULL
          AND CAST(TRIM(CAST(o.type AS CHAR)) AS BINARY) = 0x4F72646572
          AND COALESCE(CAST(o.quantity AS DECIMAL(18,4)), 0) > 0
          AND COALESCE(CAST(o.`product sales` AS DECIMAL(18,4)), 0) > 0
          AND o.sku IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(o.sku AS CHAR))) > 0
          AND o.`order id` IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(o.`order id` AS CHAR))) > 0
          AND LEFT(CAST(TRIM(CAST(o.`order id` AS CHAR)) AS BINARY), 3) <> 0x533031
          {store_filter(stores, "o")}
        """.strip()
        )
    return "\nUNION ALL\n".join(parts)


def common_ctes(tables: Sequence[str], stores: Sequence[str]) -> str:
    """复用 c13f66c 的订单过滤与三路 SKU -> MSKU 映射。"""
    return f"""
    WITH listing_fnsku AS (
        SELECT
            CAST(UPPER(REPLACE(TRIM(CAST(`店铺` AS CHAR)), '_', '-')) AS BINARY) AS store_key,
            CAST(TRIM(CAST(FNSKU AS CHAR)) AS BINARY) AS fnsku_key,
            MIN(CONVERT(TRIM(CAST(MSKU AS CHAR)) USING utf8mb4)) AS msku
        FROM lingxing.listing
        WHERE FNSKU IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(FNSKU AS CHAR))) > 0
          AND MSKU IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(MSKU AS CHAR))) > 0
          AND LEFT(CAST(TRIM(CAST(MSKU AS CHAR)) AS BINARY), 13) <> 0x{AMAZON_FOUND_HEX}
          AND SKU IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(SKU AS CHAR))) > 0
        GROUP BY store_key, fnsku_key
    ), listing_asin AS (
        SELECT
            CAST(UPPER(REPLACE(TRIM(CAST(`店铺` AS CHAR)), '_', '-')) AS BINARY) AS store_key,
            CAST(TRIM(CAST(ASIN AS CHAR)) AS BINARY) AS asin_key,
            MIN(CONVERT(TRIM(CAST(MSKU AS CHAR)) USING utf8mb4)) AS msku
        FROM lingxing.listing
        WHERE ASIN IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(ASIN AS CHAR))) > 0
          AND MSKU IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(MSKU AS CHAR))) > 0
          AND LEFT(CAST(TRIM(CAST(MSKU AS CHAR)) AS BINARY), 13) <> 0x{AMAZON_FOUND_HEX}
          AND SKU IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(SKU AS CHAR))) > 0
        GROUP BY store_key, asin_key
    ), raw_orders AS (
        {raw_order_union(tables, stores)}
    ), mapped_orders AS (
        SELECT r.order_date, r.source, r.order_id, r.sku, r.quantity, r.product_sales,
               CONVERT(CASE
                   WHEN LEFT(r.sku_key, 3) = 0x583030 THEN lf.msku
                   WHEN LEFT(r.sku_key, 2) = 0x4230
                     OR LEFT(r.sku_key, 13) = 0x{AMAZON_FOUND_HEX} THEN la.msku
                   WHEN LOCATE(0x2D, r.sku_key) > 0 THEN r.sku
                   ELSE NULL
               END USING utf8mb4) AS msku
        FROM raw_orders r
        LEFT JOIN listing_fnsku lf
          ON r.store_key = lf.store_key AND r.sku_key = lf.fnsku_key
        LEFT JOIN listing_asin la
          ON r.store_key = la.store_key
         AND (CASE WHEN LEFT(r.sku_key, 13) = 0x{AMAZON_FOUND_HEX}
                   THEN SUBSTRING(r.sku_key, 14) ELSE r.sku_key END) = la.asin_key
    ), clean_orders AS (
        SELECT order_date, source, order_id, sku, TRIM(msku) AS msku,
               TRIM(SUBSTRING_INDEX(msku, '-', 1)) AS spu,
               quantity, product_sales
        FROM mapped_orders
        WHERE msku IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(msku AS CHAR))) > 0
          AND CHAR_LENGTH(TRIM(SUBSTRING_INDEX(CAST(msku AS CHAR), '-', 1))) > 0
    )
    """


def color_system_sql(tables: Sequence[str], stores: Sequence[str]) -> str:
    """按 MSKU 求首销，再经 listing 聚合到本地 SKU。"""
    return f"""
    {common_ctes(tables, stores)},
    msku_first_sales AS (
        SELECT msku, MIN(order_date) AS first_sale_date
        FROM clean_orders
        GROUP BY msku
    ), msku_first_sale_stores AS (
        SELECT f.msku, f.first_sale_date, c.source
        FROM msku_first_sales f
        JOIN clean_orders c
          ON CAST(c.msku AS BINARY) = CAST(f.msku AS BINARY)
         AND c.order_date = f.first_sale_date
        GROUP BY f.msku, f.first_sale_date, c.source
    ), msku_recent_sales AS (
        SELECT msku,
               MAX(CASE WHEN order_date >= DATE_SUB(CURDATE(), INTERVAL 6 MONTH)
                        THEN 1 ELSE 0 END) AS has_recent_sales
        FROM clean_orders
        GROUP BY msku
    ), listing_sku_msku AS (
        SELECT DISTINCT
               CONVERT(TRIM(CAST(l.SKU AS CHAR)) USING utf8mb4) AS sku,
               CAST(TRIM(CAST(l.SKU AS CHAR)) AS BINARY) AS sku_key,
               CONVERT(TRIM(CAST(l.MSKU AS CHAR)) USING utf8mb4) AS msku,
               CAST(TRIM(CAST(l.MSKU AS CHAR)) AS BINARY) AS msku_key
        FROM lingxing.listing l
        WHERE l.SKU IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(l.SKU AS CHAR))) > 0
          AND l.MSKU IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(l.MSKU AS CHAR))) > 0
          AND LEFT(CAST(TRIM(CAST(l.MSKU AS CHAR)) AS BINARY), 13) <> 0x{AMAZON_FOUND_HEX}
          {listing_store_filter(stores, "l")}
    ), sku_opening AS (
        SELECT l.sku_key,
               MIN(l.sku) AS sku,
               MIN(f.first_sale_date) AS opening_date,
               MAX(COALESCE(r.has_recent_sales, 0)) AS has_recent_sales
        FROM listing_sku_msku l
        LEFT JOIN msku_first_sales f ON l.msku_key = CAST(f.msku AS BINARY)
        LEFT JOIN msku_recent_sales r ON l.msku_key = CAST(r.msku AS BINARY)
        GROUP BY l.sku_key
    )
    SELECT o.sku AS sku,
           GROUP_CONCAT(DISTINCT l.msku ORDER BY l.msku SEPARATOR ',') AS associated_msku,
           o.opening_date AS opening_date,
           GROUP_CONCAT(DISTINCT fs.source ORDER BY fs.source SEPARATOR ',') AS first_sale_stores,
           o.has_recent_sales AS has_recent_sales
    FROM sku_opening o
    JOIN listing_sku_msku l ON l.sku_key = o.sku_key
    LEFT JOIN msku_first_sale_stores fs
      ON l.msku_key = CAST(fs.msku AS BINARY)
     AND fs.first_sale_date = o.opening_date
    GROUP BY o.sku_key, o.sku, o.opening_date, o.has_recent_sales
    ORDER BY o.opening_date, o.sku
    """


def connect_analytics_db() -> Any:
    settings.validate_db()
    config = dict(settings.db_config)
    config.update(
        connect_timeout=60,
        read_timeout=7200,
        write_timeout=7200,
        autocommit=True,
        cursorclass=DictCursor,
        charset="utf8mb4",
    )
    return pymysql.connect(**config)


def run_analysis_query(conn: Any, sql: str) -> list[dict[str, Any]]:
    started = datetime.now()
    with conn.cursor() as cursor:
        cursor.execute("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci")
        cursor.execute("SET SESSION group_concat_max_len = 1048576")
        cursor.execute(sql)
        rows = list(cursor.fetchall())
    elapsed = (datetime.now() - started).total_seconds()
    print(f"MSKU 首销与本地 SKU 聚合完成：{len(rows):,} 个 SKU，耗时 {elapsed:.1f}s")
    return rows


def as_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def classify_opening_date(opening_date: date | None) -> str | None:
    if opening_date is None:
        return "待定"
    if opening_date <= FIRST_ROUND_CUTOFF:
        return "A2023"
    return None


def prepare_analysis_rows(rows: Iterable[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    prepared: list[dict[str, Any]] = []
    skipped = 0
    for row in rows:
        sku = clean(row.get("sku"))
        if not sku:
            skipped += 1
            continue
        opening_date = as_date(row.get("opening_date"))
        proposed = classify_opening_date(opening_date)
        if proposed is None:
            skipped += 1
            continue
        prepared.append(
            {
                "SKU": sku,
                "关联 MSKU": clean(row.get("associated_msku")),
                "SPU": extract_spu(sku),
                "开售日期": opening_date,
                "首销店铺": clean(row.get("first_sale_stores")),
                "拟打值": proposed,
                "近半年是否有销量": "是" if int(row.get("has_recent_sales") or 0) else "否",
            }
        )
    prepared.sort(
        key=lambda item: (
            0 if item["拟打值"] == "A2023" else 1,
            item["开售日期"] or date.max,
            item["SKU"],
        )
    )
    return prepared, skipped


def style_header(ws: Any, row_no: int = 1) -> None:
    for cell in ws[row_no]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws: Any, max_width: int = 60) -> None:
    for cells in ws.columns:
        letter = get_column_letter(cells[0].column)
        width = max((len(str(cell.value or "")) for cell in cells[:3000]), default=8) + 2
        ws.column_dimensions[letter].width = min(max(width, 10), max_width)


def create_dryrun_workbook(rows: Sequence[dict[str, Any]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    detail = workbook.active
    detail.title = "拟打标明细"
    detail.append(DETAIL_HEADERS)
    for row in rows:
        detail.append([row.get(header) for header in DETAIL_HEADERS])
    style_header(detail)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for cell in detail["D"][1:]:
        cell.number_format = "yyyy-mm-dd"
    if detail.max_row >= 2:
        validation = DataValidation(type="list", formula1='"A2023,待定"', allow_blank=False)
        validation.error = "拟打值只允许 A2023 或 待定"
        validation.errorTitle = "非法拟打值"
        detail.add_data_validation(validation)
        validation.add(f"F2:F{detail.max_row}")
    auto_width(detail)

    summary = workbook.create_sheet("汇总")
    summary.append(["拟打值", "SKU数"])
    counts = Counter(str(row["拟打值"]) for row in rows)
    for value in ["A2023", "待定"]:
        summary.append([value, counts[value]])
    style_header(summary, 1)

    summary.append([])
    section_row = summary.max_row + 1
    summary.append(["按 SPU 聚合分布"])
    summary.cell(section_row, 1).fill = SECTION_FILL
    summary.cell(section_row, 1).font = Font(name="Microsoft YaHei", bold=True)
    header_row = summary.max_row + 1
    summary.append(["SPU", "A2023 SKU数", "待定 SKU数", "合计"])
    by_spu: dict[str, Counter[str]] = defaultdict(Counter)
    for row in rows:
        by_spu[str(row["SPU"])][str(row["拟打值"])] += 1
    for spu in sorted(by_spu):
        a_count = by_spu[spu]["A2023"]
        pending_count = by_spu[spu]["待定"]
        summary.append([spu, a_count, pending_count, a_count + pending_count])
    style_header(summary, header_row)
    summary.freeze_panes = "A2"
    auto_width(summary)
    workbook.save(output)


def load_review_file(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise FileNotFoundError(f"审阅清单不存在：{path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "拟打标明细" not in workbook.sheetnames:
            raise ValueError("审阅清单缺少 sheet：拟打标明细")
        sheet = workbook["拟打标明细"]
        headers = {
            clean(cell.value): index
            for index, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)))
            if clean(cell.value)
        }
        missing = [header for header in ("SKU", "拟打值") if header not in headers]
        if missing:
            raise ValueError(f"审阅清单缺少列：{','.join(missing)}")

        targets: list[dict[str, str]] = []
        seen: set[str] = set()
        for excel_row_no, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            sku = clean(cells[headers["SKU"]] if headers["SKU"] < len(cells) else "")
            value = clean(cells[headers["拟打值"]] if headers["拟打值"] < len(cells) else "")
            if not sku and not value:
                continue
            if not sku:
                raise ValueError(f"审阅清单第 {excel_row_no} 行 SKU 为空")
            if value not in TARGET_VALUES:
                raise ValueError(
                    f"审阅清单第 {excel_row_no} 行拟打值非法：{value!r}；只允许 A2023/待定"
                )
            if sku in seen:
                raise ValueError(f"审阅清单存在重复 SKU：{sku}")
            seen.add(sku)
            targets.append({"SKU": sku, "拟打值": value})
        return targets
    finally:
        workbook.close()


def is_beijing_low_peak(now: datetime | None = None) -> bool:
    current = (now or datetime.now(BEIJING_TZ)).astimezone(BEIJING_TZ).time()
    return LOW_PEAK_START <= current < LOW_PEAK_END


def validate_field_id(value: str) -> str:
    field_id = clean(value)
    if field_id and not field_id.isdigit():
        raise ValueError(f"颜色体系字段 ID 非法：{field_id!r}")
    return field_id


def field_ids_from_fields(fields: Any) -> set[str]:
    if isinstance(fields, str):
        try:
            fields = json.loads(fields)
        except json.JSONDecodeError:
            return set()
    if not isinstance(fields, list):
        return set()
    return {
        clean(item.get("id"))
        for item in fields
        if isinstance(item, dict)
        and clean(item.get("name")) == COLOR_FIELD_NAME
        and clean(item.get("id"))
    }


def one_field_id(field_ids: set[str], source: str) -> str:
    if len(field_ids) > 1:
        raise RuntimeError(f"{source} 中发现多个“{COLOR_FIELD_NAME}”字段 ID：{sorted(field_ids)}")
    return next(iter(field_ids), "")


def field_id_from_snapshot(db: Database) -> str:
    try:
        row = db.fetch_one(
            """
            SELECT custom_fields_json
            FROM lxpm_product_category_snapshot
            WHERE custom_fields_json LIKE %s
            ORDER BY synced_at DESC
            LIMIT 1
            """,
            (f'%"name":"{COLOR_FIELD_NAME}"%',),
        )
    except Exception as exc:  # noqa: BLE001
        print(f"警告：无法从产品快照发现颜色体系字段 ID：{exc}")
        return ""
    return one_field_id(field_ids_from_fields((row or {}).get("custom_fields_json")), "产品快照")


async def discover_field_id_from_live(
    client: LingxingClient,
    token: str,
    skus: Sequence[str],
) -> tuple[str, str]:
    found: set[str] = set()
    current_token = token
    sample = list(skus[:500])
    for start in range(0, len(sample), 100):
        live_map, current_token = await fetch_live_products(
            client,
            current_token,
            sample[start : start + 100],
            max_retries=RATE_LIMIT_RETRIES,
            batch_size=100,
            delay_seconds=0.5,
            retry_base_seconds=RATE_LIMIT_WAIT_SECONDS,
            retry_max_seconds=RATE_LIMIT_WAIT_SECONDS,
        )
        for product in live_map.values():
            found.update(field_ids_from_fields(extract_custom_fields(product)))
        if found:
            break
    return one_field_id(found, "领星实时产品详情"), current_token


def failure_record(
    sku: str,
    value: str,
    stage: str,
    error: Any,
    response: dict[str, Any] | None = None,
) -> dict[str, str]:
    response = response or {}
    message = clean(response.get("msg") or response.get("message")) or clean(error)
    return {
        "SKU": sku,
        "拟打值": value,
        "失败阶段": stage,
        "错误码": clean(response.get("code")),
        "错误信息": message[:2000],
        "request_id": clean(response.get("request_id")),
    }


def response_succeeded(response: dict[str, Any] | None) -> bool:
    return str((response or {}).get("code")) == "0"


def create_failure_workbook(rows: Sequence[dict[str, str]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "失败清单"
    sheet.append(FAILURE_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in FAILURE_HEADERS])
    style_header(sheet)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    auto_width(sheet, max_width=80)
    workbook.save(output)


async def apply_review_file(
    args: argparse.Namespace,
    *,
    write_values: set[str] | None = None,
    failure_prefix: str = "color_system_tagging_failures",
) -> int:
    if not args.review_file:
        raise ValueError("--apply 必须同时指定 --review-file，禁止直接生成后写入")
    if not args.allow_outside_low_peak and not is_beijing_low_peak():
        now_text = datetime.now(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")
        raise RuntimeError(
            f"当前北京时间 {now_text} 不在低峰窗口 00:00-06:00；"
            "请在凌晨执行，或人工确认后使用 --allow-outside-low-peak"
        )

    review_path = Path(args.review_file).expanduser().resolve()
    review_targets = load_review_file(review_path)
    if write_values is not None:
        invalid_write_values = set(write_values) - TARGET_VALUES
        if invalid_write_values:
            raise ValueError(f"非法写入值过滤：{sorted(invalid_write_values)}")
        targets = [item for item in review_targets if item["拟打值"] in write_values]
    else:
        targets = review_targets
    preserved_count = len(review_targets) - len(targets)
    print("===== 颜色体系审阅清单写入 =====")
    print(f"审阅清单：{review_path}")
    print(f"处理 SKU：{len(review_targets):,}")
    print(f"计划写入：{len(targets):,}；按规则不写：{preserved_count:,}")
    print("安全策略：实时产品详情 + 完整自定义字段合并 + product/set")
    print("code=103：固定等待 30 秒，最多重试 3 次")
    if not targets:
        print("没有需要写入的行")
        print(
            f"执行统计：处理={len(review_targets):,} 成功=0 失败=0 "
            f"跳过={preserved_count:,}"
        )
        return 0

    db = Database()
    client = LingxingClient(db=db, enable_api_log=True)
    failures: list[dict[str, str]] = []
    stats: Counter[str] = Counter(
        processed=len(review_targets),
        skipped=preserved_count,
    )
    try:
        token = (await client.generate_token()).token
        field_id = validate_field_id(args.field_id or os.getenv(COLOR_FIELD_ID_ENV, ""))
        if not field_id:
            field_id = field_id_from_snapshot(db)
        if not field_id:
            field_id, token = await discover_field_id_from_live(
                client,
                token,
                [item["SKU"] for item in targets],
            )
        if not field_id:
            raise RuntimeError(
                f"无法发现“{COLOR_FIELD_NAME}”字段 ID；请通过 --field-id 或 "
                f"{COLOR_FIELD_ID_ENV} 显式提供"
            )
        print(f"颜色体系字段 ID：{field_id}")

        batch_size = max(1, min(int(args.batch_size), 100))
        for start in range(0, len(targets), batch_size):
            batch = targets[start : start + batch_size]
            try:
                live_map, token = await fetch_live_products(
                    client,
                    token,
                    [item["SKU"] for item in batch],
                    max_retries=RATE_LIMIT_RETRIES,
                    batch_size=batch_size,
                    delay_seconds=0.5,
                    retry_base_seconds=RATE_LIMIT_WAIT_SECONDS,
                    retry_max_seconds=RATE_LIMIT_WAIT_SECONDS,
                )
            except Exception as exc:  # noqa: BLE001
                for item in batch:
                    failures.append(
                        failure_record(item["SKU"], item["拟打值"], "写前实时读取", exc)
                    )
                    stats["failed"] += 1
                continue

            for item in batch:
                sku = item["SKU"]
                value = item["拟打值"]
                target_field = {"id": field_id, "name": COLOR_FIELD_NAME, "val": value}
                live_product = live_map.get(sku)
                try:
                    if not live_product:
                        raise RuntimeError("写前实时详情未返回该 SKU")
                    discovered_ids = field_ids_from_fields(extract_custom_fields(live_product))
                    if discovered_ids and discovered_ids != {field_id}:
                        raise RuntimeError(
                            f"实时详情字段 ID 与配置不一致：configured={field_id}, "
                            f"live={sorted(discovered_ids)}"
                        )
                    if target_fields_match(extract_custom_fields(live_product), [target_field]):
                        stats["skipped"] += 1
                        continue

                    body, _guard_meta = build_guarded_product_set_body(
                        live_product,
                        sku=sku,
                        target_custom_fields=[target_field],
                        preserve_current_category=True,
                    )
                    response, token = await request_with_retry(
                        client,
                        token,
                        PRODUCT_SET_API,
                        body,
                        max_retries=RATE_LIMIT_RETRIES,
                        retry_base_seconds=RATE_LIMIT_WAIT_SECONDS,
                        retry_max_seconds=RATE_LIMIT_WAIT_SECONDS,
                    )
                    if response_succeeded(response):
                        stats["success"] += 1
                    else:
                        failures.append(
                            failure_record(sku, value, "product/set", response, response=response)
                        )
                        stats["failed"] += 1
                except Exception as exc:  # noqa: BLE001
                    failures.append(failure_record(sku, value, "安全合并写入", exc))
                    stats["failed"] += 1

                done = stats["success"] + stats["failed"] + stats["skipped"]
                if done % 100 == 0 or done == len(review_targets):
                    print(
                        f"写入进度：{done:,}/{len(review_targets):,} "
                        f"成功={stats['success']:,} 失败={stats['failed']:,} "
                        f"跳过={stats['skipped']:,}"
                    )
                if args.delay > 0:
                    await asyncio.sleep(args.delay)
    finally:
        await client.aclose()
        db.close()

    failure_path: Path | None = None
    if failures:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        failure_path = OUTPUT_DIR / (
            f"{failure_prefix}_{datetime.now(BEIJING_TZ):%Y%m%d_%H%M%S}.xlsx"
        )
        create_failure_workbook(failures, failure_path)

    print(
        "执行统计："
        f"处理={stats['processed']:,} 成功={stats['success']:,} "
        f"失败={stats['failed']:,} 跳过={stats['skipped']:,}"
    )
    if failure_path:
        print(f"失败清单：{failure_path}")
    return 1 if failures else 0


def run_dry_run(args: argparse.Namespace) -> int:
    years = parse_years(args.years)
    stores = parse_stores(args.stores)
    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else OUTPUT_DIR / f"color_system_tagging_dryrun_{datetime.now(BEIJING_TZ):%Y%m%d}.xlsx"
    )

    print("===== 颜色体系批量打标 dry-run（第一轮） =====")
    print(f"年份：{years}")
    print(f"店铺：{'ALL' if not stores else ','.join(stores)}")
    print(f"A2023 截止日期：{FIRST_ROUND_CUTOFF:%Y-%m-%d}")
    conn = connect_analytics_db()
    try:
        tables = resolve_tables(conn, years)
        raw_rows = run_analysis_query(conn, color_system_sql(tables, stores))
    finally:
        conn.close()

    targets, skipped = prepare_analysis_rows(raw_rows)
    create_dryrun_workbook(targets, output)
    counts = Counter(row["拟打值"] for row in targets)
    for row in targets[: max(0, args.show)]:
        print(
            f"{row['SKU']} | {row['关联 MSKU']} | {row['开售日期'] or ''} "
            f"-> {row['拟打值']}"
        )
    print(
        "执行统计："
        f"处理={len(raw_rows):,} 成功=0 失败=0 跳过={skipped:,} "
        f"拟打A2023={counts['A2023']:,} 拟打待定={counts['待定']:,}"
    )
    print(f"dry-run 清单：{output}")
    return 0


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if args.apply:
        return asyncio.run(apply_review_file(args))
    if args.review_file:
        raise ValueError("--review-file 只允许与 --apply 一起使用")
    return run_dry_run(args)


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n用户中断")
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001
        print(f"执行失败：{type(exc).__name__}: {exc}")
        raise
