#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系只读诊断：使用内嵌颜色表匹配 SKU 颜色代码与产品品名。

安全边界：
1. 只读取 MySQL 与产品快照。
2. 只生成 Excel，不包含 --apply，不调用领星 product/set。
3. 主范围为颜色体系为空的全部本地产品 SKU。
4. 已有 A2023 SKU 仅用于回测，不修改。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from lx_product_m.color_system_mapping_data import (
    COLOR_MAPPING_ROWS,
    COLOR_MAPPING_ROW_COUNT,
    COLOR_MAPPING_SOURCE,
)
from lx_product_m.db import Database
from lx_product_m.services.product_write_guard import clean
from scripts.color_system_tagging_analysis import common_ctes, connect_analytics_db

COLOR_FIELD_ID = "207722905719915521"
COLOR_FIELD_NAME = "颜色体系"
SYSTEMS = ("A2023", "B2024")
STYLE_TOKENS = {"SHORT", "LONG", "TALL", "PETITE"}
GENERIC_COLOR_WORDS = {
    "净色", "纯色", "撞色", "拼色", "花色", "多色", "有色", "无色",
}
OUTPUT_DIR = ROOT / "reports_analysis"

DETAIL_HEADERS = [
    "SKU", "产品名称", "SPU", "分类路径", "当前颜色体系",
    "SKU结构", "是否多色套装", "件数", "版型", "颜色代码", "尺码",
    "品名识别颜色", "A2023主映射", "A2023次级映射",
    "B2024主映射", "B2024次级映射",
    "拟打颜色体系", "置信度", "判定类型", "匹配方式", "判定原因",
    "历史是否有销量", "首次销售日期", "最近销售日期",
    "Listing记录数", "最早Listing创建时间", "处理优先级",
    "是否允许未来自动写入",
]
BACKTEST_HEADERS = DETAIL_HEADERS + ["回测是否命中A2023"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


@dataclass(frozen=True)
class ColorEntry:
    source_position: int
    sequence: int
    code: str
    english: str
    chinese: str
    pantone: str
    system: str
    is_primary: bool


@dataclass(frozen=True)
class SkuParseResult:
    sku: str
    spu: str
    structure: str
    color_codes: tuple[str, ...]
    size: str
    style: str
    is_bundle: bool
    pieces: str
    error: str = ""


@dataclass(frozen=True)
class MatchHit:
    entry: ColorEntry
    label: str
    matched_text: str
    match_kind: str
    matched_length: int


@dataclass(frozen=True)
class Decision:
    proposed: str
    confidence: str
    decision_type: str
    match_method: str
    reason: str
    matched_colors: tuple[str, ...]


@dataclass
class MappingIndex:
    entries: list[ColorEntry]
    by_system_code: dict[tuple[str, str], list[ColorEntry]]
    primary: dict[tuple[str, str], ColorEntry]
    codes: set[str]
    codes_by_length: tuple[str, ...]
    systems_by_code: dict[str, set[str]]
    variant_to_labels: dict[str, set[str]]
    global_pattern: re.Pattern[str] | None


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="颜色体系只读诊断：内嵌颜色表 + SKU颜色代码 + 品名颜色"
    )
    parser.add_argument(
        "--output",
        default="",
        help="输出 Excel；默认 reports_analysis/color_system_name_code_analysis_时间.xlsx",
    )
    parser.add_argument(
        "--skip-sales",
        action="store_true",
        help="跳过历史订单扫描；仅用于快速试跑，优先级中的销量将标记为未知",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="仅分析前 N 个空标签 SKU，默认0表示全量；用于功能测试",
    )
    parser.add_argument("--show", type=int, default=20, help="控制台展示高置信度候选数量")
    return parser.parse_args(argv)


def normalize_text(value: Any) -> str:
    return unicodedata.normalize("NFKC", clean(value))


def normalize_code(value: Any) -> str:
    return normalize_text(value).upper().replace(" ", "")


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def split_color_labels(chinese: str) -> list[str]:
    """把“米色（米黄）”“橙色/橘色”拆成可匹配标签。"""
    value = compact_text(chinese)
    if not value:
        return []
    expanded = re.sub(r"[（(]([^）)]+)[）)]", r"/\1", value)
    parts = re.split(r"[/、,，;；]+", expanded)
    output: list[str] = []
    for part in parts:
        label = part.strip(" -_")
        if label and label not in output:
            output.append(label)
    return output


def color_variants(label: str) -> list[tuple[str, str]]:
    value = compact_text(label)
    if not value or value in GENERIC_COLOR_WORDS:
        return []
    variants = [(value, "完整颜色")]
    if value.endswith("色") and len(value) > 1:
        stripped = value[:-1]
        if stripped and stripped not in GENERIC_COLOR_WORDS:
            variants.append((stripped, "缺少色字"))
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    for variant, kind in variants:
        if variant not in seen:
            seen.add(variant)
            out.append((variant, kind))
    return out


def build_mapping_index(
    rows: Sequence[tuple[int, str, str, str, str, str]] = COLOR_MAPPING_ROWS,
) -> MappingIndex:
    by_system_code: dict[tuple[str, str], list[ColorEntry]] = defaultdict(list)
    entries: list[ColorEntry] = []

    for position, raw in enumerate(rows, start=1):
        sequence, code, english, chinese, pantone, system = raw
        code_norm = normalize_code(code)
        system_norm = normalize_text(system)
        key = (system_norm, code_norm)
        entry = ColorEntry(
            source_position=position,
            sequence=int(sequence),
            code=code_norm,
            english=normalize_text(english),
            chinese=normalize_text(chinese),
            pantone=normalize_text(pantone),
            system=system_norm,
            is_primary=not by_system_code[key],
        )
        by_system_code[key].append(entry)
        entries.append(entry)

    primary = {key: values[0] for key, values in by_system_code.items()}
    codes = {entry.code for entry in entries if entry.code}
    systems_by_code: dict[str, set[str]] = defaultdict(set)
    variant_to_labels: dict[str, set[str]] = defaultdict(set)

    for entry in entries:
        systems_by_code[entry.code].add(entry.system)
        for label in split_color_labels(entry.chinese):
            for variant, _ in color_variants(label):
                if len(variant) >= 2:
                    variant_to_labels[variant].add(label)

    alternatives = sorted(variant_to_labels, key=lambda x: (-len(x), x))
    global_pattern = (
        re.compile(r"(?=(" + "|".join(re.escape(x) for x in alternatives) + r"))")
        if alternatives
        else None
    )

    return MappingIndex(
        entries=entries,
        by_system_code=dict(by_system_code),
        primary=primary,
        codes=codes,
        codes_by_length=tuple(sorted(codes, key=lambda x: (-len(x), x))),
        systems_by_code={k: set(v) for k, v in systems_by_code.items()},
        variant_to_labels={k: set(v) for k, v in variant_to_labels.items()},
        global_pattern=global_pattern,
    )


def parse_custom_fields(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return []
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def field_value(item: dict[str, Any]) -> str:
    for key in ("val_text", "val", "value", "field_value"):
        if key in item and item.get(key) is not None:
            return clean(item.get(key))
    return ""


def extract_color_system(value: Any) -> str:
    values: list[str] = []
    for item in parse_custom_fields(value):
        field_id = clean(item.get("id"))
        name = clean(item.get("name"))
        if field_id != COLOR_FIELD_ID and name != COLOR_FIELD_NAME:
            continue
        current = field_value(item)
        if current and current not in values:
            values.append(current)
    return "|".join(values)


def parse_sku(sku: str, index: MappingIndex) -> SkuParseResult:
    raw = normalize_text(sku).upper().replace("_", "-")
    raw = re.sub(r"-+", "-", raw).strip("-")
    if not raw:
        return SkuParseResult("", "", "无法解析", (), "", "", False, "", "SKU为空")

    parts = [x for x in raw.split("-") if x]
    spu = parts[0] if parts else ""
    if len(parts) < 2:
        return SkuParseResult(raw, spu, "无法解析", (), "", "", False, "", "SKU缺少颜色与尺码段")

    body = parts[1:]
    style = ""
    if body and body[0] in STYLE_TOKENS:
        style = body.pop(0)

    pcs_index = next(
        (i for i, part in enumerate(body) if re.fullmatch(r"\d+(?:PCS|PSC)", part, re.I)),
        None,
    )
    if pcs_index is not None:
        pieces = body[pcs_index]
        after = body[pcs_index + 1 :]
        if len(after) < 2:
            return SkuParseResult(
                raw, spu, "PCS多色套装", (), "", style, True, pieces,
                "PCS后缺少完整颜色与尺码",
            )
        size = after[-1]
        color_parts = after[:-1]
        valid = tuple(part for part in color_parts if part in index.codes)
        unknown = [part for part in color_parts if part not in index.codes]
        if unknown:
            return SkuParseResult(
                raw, spu, "PCS多色套装", valid, size, style, True, pieces,
                "存在未知颜色段：" + ",".join(unknown),
            )
        if not valid:
            return SkuParseResult(
                raw, spu, "PCS多色套装", (), size, style, True, pieces,
                "未识别到颜色代码",
            )
        return SkuParseResult(raw, spu, "PCS多色套装", valid, size, style, True, pieces)

    if not body:
        return SkuParseResult(raw, spu, "无法解析", (), "", style, False, "", "缺少颜色段")

    first = body[0]
    if first in index.codes:
        size = "-".join(body[1:])
        if not size:
            return SkuParseResult(
                raw, spu, "常规", (first,), "", style, False, "", "缺少尺码段",
            )
        extra_codes = [part for part in body[1:-1] if part in index.codes]
        if extra_codes:
            return SkuParseResult(
                raw, spu, "非PCS多颜色", (first, *extra_codes), body[-1],
                style, True, "", "发现多个颜色代码但无PCS标记",
            )
        return SkuParseResult(raw, spu, "常规", (first,), size, style, False, "")

    glued = [
        code for code in index.codes_by_length
        if first.startswith(code) and len(first) > len(code)
    ]
    if glued:
        max_len = len(glued[0])
        best = [code for code in glued if len(code) == max_len]
        if len(best) > 1:
            return SkuParseResult(
                raw, spu, "颜色尺码粘连", tuple(best), "", style, False, "",
                "颜色尺码粘连存在同长度代码歧义",
            )
        code = best[0]
        size = first[len(code):]
        trailing = "-".join(body[1:])
        if trailing:
            size = f"{size}-{trailing}" if size else trailing
        return SkuParseResult(raw, spu, "颜色尺码粘连", (code,), size, style, False, "")

    scan_codes = [part for part in body[:-1] if part in index.codes]
    if len(scan_codes) == 1:
        return SkuParseResult(
            raw, spu, "位置异常但可识别", (scan_codes[0],), body[-1],
            style, False, "", "颜色代码未处于常规位置",
        )
    if len(scan_codes) > 1:
        return SkuParseResult(
            raw, spu, "非PCS多颜色", tuple(scan_codes), body[-1],
            style, True, "", "多个颜色代码且无PCS标记",
        )

    return SkuParseResult(
        raw, spu, "无法解析", (), body[-1] if body else "", style, False, "",
        "颜色代码不在内嵌映射表",
    )


def single_char_variant_found(name: str, variant: str) -> bool:
    return bool(re.search(re.escape(variant) + r"(?=$|[-_/（）()\[\]\s\d])", name))


def match_entry(name: str, entry: ColorEntry) -> MatchHit | None:
    normalized_name = compact_text(name)
    best: MatchHit | None = None
    for label in split_color_labels(entry.chinese):
        for variant, kind in color_variants(label):
            found = variant in normalized_name if len(variant) >= 2 else single_char_variant_found(normalized_name, variant)
            if not found:
                continue
            hit = MatchHit(entry, label, variant, kind, len(variant))
            if best is None or (hit.matched_length, hit.match_kind == "完整颜色") > (
                best.matched_length, best.match_kind == "完整颜色"
            ):
                best = hit
    return best


def fuzzy_entry_match(name: str, entry: ColorEntry) -> MatchHit | None:
    """仅做同长度、单字替换候选；不直接进入自动写入。"""
    normalized_name = compact_text(name)
    best: MatchHit | None = None
    for label in split_color_labels(entry.chinese):
        candidate = compact_text(label)
        if candidate in GENERIC_COLOR_WORDS or len(candidate) < 3 or len(candidate) > 8:
            continue
        length = len(candidate)
        for start in range(0, max(0, len(normalized_name) - length + 1)):
            window = normalized_name[start:start + length]
            distance = sum(a != b for a, b in zip(window, candidate))
            if distance != 1:
                continue
            hit = MatchHit(entry, label, window, "单字模糊候选", length)
            if best is None or hit.matched_length > best.matched_length:
                best = hit
    return best


def global_name_colors(name: str, index: MappingIndex) -> tuple[str, ...]:
    text = compact_text(name)
    if not text or index.global_pattern is None:
        return ()
    raw: list[tuple[int, int, str]] = []
    for match in index.global_pattern.finditer(text):
        value = match.group(1)
        raw.append((match.start(1), match.start(1) + len(value), value))
    raw.sort(key=lambda x: (x[0], -(x[1] - x[0])))
    selected: list[tuple[int, int, str]] = []
    for item in raw:
        if any(item[0] >= existing[0] and item[1] <= existing[1] for existing in selected):
            continue
        selected.append(item)
    labels: list[str] = []
    for _, _, variant in selected:
        for label in sorted(index.variant_to_labels.get(variant, {variant})):
            if label not in labels:
                labels.append(label)
    return tuple(labels)


def mapping_display(index: MappingIndex, system: str, code: str) -> tuple[str, str]:
    entries = index.by_system_code.get((system, code), [])
    if not entries:
        return "", ""
    primary = entries[0].chinese
    secondary = "；".join(entry.chinese for entry in entries[1:])
    return primary, secondary


def select_longest_hits(hits: Iterable[MatchHit]) -> list[MatchHit]:
    values = list(hits)
    if not values:
        return []
    max_length = max(hit.matched_length for hit in values)
    return [hit for hit in values if hit.matched_length == max_length]


def decide_single_code(code: str, product_name: str, index: MappingIndex) -> Decision:
    code = normalize_code(code)
    entries = [
        entry
        for system in SYSTEMS
        for entry in index.by_system_code.get((system, code), [])
    ]
    if not entries:
        return Decision("", "低", "SKU结构异常", "无映射", "颜色代码不存在于内嵌颜色表", ())

    exact_hits = select_longest_hits(
        hit for entry in entries if (hit := match_entry(product_name, entry)) is not None
    )
    exact_systems = {hit.entry.system for hit in exact_hits}
    matched_colors = tuple(dict.fromkeys(hit.label for hit in exact_hits))

    if len(exact_systems) == 1:
        system = next(iter(exact_systems))
        primary_hit = any(hit.entry.is_primary for hit in exact_hits)
        if primary_hit:
            return Decision(
                system, "高", "精确唯一匹配", "代码+品名主映射",
                f"颜色代码 {code} 与品名颜色仅命中 {system} 主映射，另一体系不匹配",
                matched_colors,
            )
        return Decision(
            system, "中", "次级映射匹配", "代码+品名次级映射",
            f"颜色代码 {code} 与品名颜色仅命中 {system} 次级历史映射；按原表第一条优先，需审阅",
            matched_colors,
        )

    if len(exact_systems) > 1:
        return Decision(
            "", "低", "A/B同时匹配", "代码+品名同时命中",
            f"颜色代码 {code} 和品名颜色在 A2023、B2024 中都成立，无法区分",
            matched_colors,
        )

    fuzzy_hits = select_longest_hits(
        hit for entry in entries if (hit := fuzzy_entry_match(product_name, entry)) is not None
    )
    fuzzy_systems = {hit.entry.system for hit in fuzzy_hits}
    if len(fuzzy_systems) == 1:
        system = next(iter(fuzzy_systems))
        colors = tuple(dict.fromkeys(hit.label for hit in fuzzy_hits))
        return Decision(
            system, "低", "模糊匹配待审阅", "代码+品名单字模糊",
            f"颜色代码 {code} 在 {system} 出现单字差异的品名颜色候选，不自动确认",
            colors,
        )
    if len(fuzzy_systems) > 1:
        colors = tuple(dict.fromkeys(hit.label for hit in fuzzy_hits))
        return Decision(
            "", "低", "模糊匹配歧义", "代码+品名单字模糊",
            f"颜色代码 {code} 的模糊品名候选同时指向 A2023、B2024",
            colors,
        )

    known_name_colors = global_name_colors(product_name, index)
    code_systems = index.systems_by_code.get(code, set())
    if known_name_colors:
        return Decision(
            "", "低", "代码与品名冲突", "全局颜色冲突",
            f"品名识别到颜色 {'、'.join(known_name_colors)}，但均不符合颜色代码 {code} 的A/B映射",
            known_name_colors,
        )

    if len(code_systems) == 1:
        system = next(iter(code_systems))
        return Decision(
            system, "中", "独占代码待审阅", "仅颜色代码",
            f"颜色代码 {code} 仅存在于 {system}，但品名未识别到对应颜色",
            (),
        )

    return Decision(
        "", "低", "缺少品名颜色证据", "重叠代码无品名命中",
        f"颜色代码 {code} 同时存在于 A2023、B2024，品名未提供可区分颜色",
        (),
    )


def decide_bundle(codes: Sequence[str], product_name: str, index: MappingIndex) -> Decision:
    if not codes:
        return Decision("", "低", "多色套装异常", "PCS解析", "未识别到套装颜色代码", ())

    possible: set[str] | None = None
    child_decisions: list[Decision] = []
    for code in codes:
        systems = index.systems_by_code.get(code, set())
        possible = set(systems) if possible is None else possible & systems
        child_decisions.append(decide_single_code(code, product_name, index))

    unique_proposals = {
        decision.proposed
        for decision in child_decisions
        if decision.proposed in SYSTEMS
    }
    conflicts = [
        (code, decision)
        for code, decision in zip(codes, child_decisions)
        if decision.decision_type in {
            "代码与品名冲突", "SKU结构异常", "A/B同时匹配", "模糊匹配歧义",
        }
    ]
    matched_colors = tuple(
        dict.fromkeys(
            color
            for decision in child_decisions
            for color in decision.matched_colors
        )
    )

    if conflicts:
        return Decision(
            "", "低", "多色套装冲突", "PCS多代码综合",
            "套装中至少一个颜色代码存在冲突或歧义：" +
            "；".join(f"{code}:{decision.decision_type}" for code, decision in conflicts),
            matched_colors,
        )

    if len(unique_proposals) == 1:
        system = next(iter(unique_proposals))
        if all(not decision.proposed or decision.proposed == system for decision in child_decisions):
            return Decision(
                system, "中", "多色套装同向候选", "PCS多代码综合",
                f"套装颜色代码的可用证据一致指向 {system}；多色套装仍需人工审阅",
                matched_colors,
            )

    if possible and len(possible) == 1:
        system = next(iter(possible))
        return Decision(
            system, "中", "多色套装独占体系", "PCS代码体系交集",
            f"套装全部颜色代码仅能共同存在于 {system}；品名未完整验证，不自动确认",
            matched_colors,
        )

    if len(unique_proposals) > 1 or not possible:
        return Decision(
            "", "低", "多色套装跨体系", "PCS多代码综合",
            "套装颜色代码的体系证据相互冲突或不存在共同体系",
            matched_colors,
        )

    return Decision(
        "", "低", "多色套装待定", "PCS多代码综合",
        "套装颜色代码同时适用于多个体系，且品名证据不足",
        matched_colors,
    )


def discover_order_tables(conn: Any) -> list[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema='sy_order'
              AND table_name REGEXP '^total_order_[0-9]{4}$'
            ORDER BY table_name
            """
        )
        rows = list(cur.fetchall())
    return [f"sy_order.{row['table_name']}" for row in rows]


def load_sales_stats(skip_sales: bool) -> tuple[dict[str, dict[str, Any]], str]:
    if skip_sales:
        return {}, "已跳过"
    conn = connect_analytics_db()
    try:
        tables = discover_order_tables(conn)
        if not tables:
            return {}, "未发现订单年度表"
        print("历史销量扫描表：" + ", ".join(tables))
        sql = f"""
{common_ctes(tables, [])},
sold_msku AS (
  SELECT CAST(TRIM(CAST(msku AS CHAR)) AS BINARY) msku_key,
         MIN(order_date) first_sale_date,
         MAX(order_date) last_sale_date
  FROM clean_orders
  GROUP BY CAST(TRIM(CAST(msku AS CHAR)) AS BINARY)
), listing_sku_msku AS (
  SELECT DISTINCT
         CAST(TRIM(CAST(l.SKU AS CHAR)) AS BINARY) sku_key,
         CONVERT(TRIM(CAST(l.SKU AS CHAR)) USING utf8mb4) sku,
         CAST(TRIM(CAST(l.MSKU AS CHAR)) AS BINARY) msku_key
  FROM lingxing.listing l
  WHERE l.SKU IS NOT NULL
    AND CHAR_LENGTH(TRIM(CAST(l.SKU AS CHAR))) > 0
    AND l.MSKU IS NOT NULL
    AND CHAR_LENGTH(TRIM(CAST(l.MSKU AS CHAR))) > 0
)
SELECT MIN(l.sku) sku,
       1 has_sales,
       MIN(s.first_sale_date) first_sale_date,
       MAX(s.last_sale_date) last_sale_date,
       COUNT(*) sold_msku_count
FROM listing_sku_msku l
JOIN sold_msku s ON l.msku_key=s.msku_key
GROUP BY l.sku_key
"""
        started = datetime.now()
        with conn.cursor() as cur:
            cur.execute("SET NAMES utf8mb4 COLLATE utf8mb4_unicode_ci")
            cur.execute(sql)
            rows = list(cur.fetchall())
        print(
            f"历史销量聚合完成：{len(rows):,} 个本地SKU，"
            f"耗时 {(datetime.now() - started).total_seconds():.1f}s"
        )
        return {
            clean(row.get("sku")): {
                "has_sales": bool(row.get("has_sales")),
                "first_sale_date": row.get("first_sale_date"),
                "last_sale_date": row.get("last_sale_date"),
                "sold_msku_count": int(row.get("sold_msku_count") or 0),
            }
            for row in rows
            if clean(row.get("sku"))
        }, "成功：" + ",".join(table.rsplit("_", 1)[-1] for table in tables)
    except Exception as exc:
        print(f"警告：历史销量扫描失败，将继续生成诊断：{type(exc).__name__}: {exc}")
        return {}, f"失败：{type(exc).__name__}: {exc}"
    finally:
        conn.close()


def load_listing_stats(db: Database) -> dict[str, dict[str, Any]]:
    rows = db.fetch_all(
        """
        SELECT MIN(CONVERT(TRIM(CAST(`SKU` AS CHAR)) USING utf8mb4)) sku,
               COUNT(*) listing_count,
               MIN(`创建时间`) earliest_listing_create_time
        FROM lingxing.listing
        WHERE `SKU` IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(`SKU` AS CHAR))) > 0
        GROUP BY CAST(TRIM(CAST(`SKU` AS CHAR)) AS BINARY)
        """
    )
    return {
        clean(row.get("sku")): {
            "listing_count": int(row.get("listing_count") or 0),
            "earliest_listing_create_time": row.get("earliest_listing_create_time"),
        }
        for row in rows
        if clean(row.get("sku"))
    }


def load_snapshot_rows(db: Database) -> list[dict[str, Any]]:
    return db.fetch_all(
        """
        SELECT sku, product_name, spu, category_path, custom_fields_json
        FROM lxpm_product_category_snapshot
        WHERE sku IS NOT NULL
          AND CHAR_LENGTH(TRIM(CAST(sku AS CHAR))) > 0
        ORDER BY sku
        """
    )


def display_date(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value
    return clean(value)


def priority_for(
    sku: str,
    sales_stats: dict[str, dict[str, Any]],
    listing_stats: dict[str, dict[str, Any]],
    sales_status: str,
) -> tuple[str, str, Any, Any, int, Any]:
    sales = sales_stats.get(sku, {})
    listing = listing_stats.get(sku, {})
    has_sales = bool(sales.get("has_sales"))
    listing_count = int(listing.get("listing_count") or 0)

    if sales_status.startswith("失败") or sales_status in {"已跳过", "未发现订单年度表"}:
        has_sales_text = "未知"
        priority = "P2-有Listing/销量未知" if listing_count else "P3-无Listing/销量未知"
    elif has_sales:
        has_sales_text = "是"
        priority = "P1-有销量"
    elif listing_count:
        has_sales_text = "否"
        priority = "P2-无销量有Listing"
    else:
        has_sales_text = "否"
        priority = "P3-无销量无Listing"

    return (
        priority,
        has_sales_text,
        display_date(sales.get("first_sale_date")),
        display_date(sales.get("last_sale_date")),
        listing_count,
        display_date(listing.get("earliest_listing_create_time")),
    )


def analyze_product(
    source: dict[str, Any],
    index: MappingIndex,
    sales_stats: dict[str, dict[str, Any]],
    listing_stats: dict[str, dict[str, Any]],
    sales_status: str,
) -> dict[str, Any]:
    sku = clean(source.get("sku"))
    product_name = clean(source.get("product_name"))
    current = extract_color_system(source.get("custom_fields_json"))
    parsed = parse_sku(sku, index)
    name_colors = global_name_colors(product_name, index)

    if parsed.error and not parsed.color_codes:
        decision = Decision("", "低", "SKU结构异常", "SKU解析", parsed.error, name_colors)
    elif parsed.is_bundle or len(parsed.color_codes) > 1:
        decision = decide_bundle(parsed.color_codes, product_name, index)
        if parsed.error:
            decision = Decision(
                decision.proposed, "低", "多色套装异常", decision.match_method,
                parsed.error + "；" + decision.reason, decision.matched_colors,
            )
    else:
        decision = decide_single_code(parsed.color_codes[0], product_name, index)

    mapping_values: dict[str, list[str]] = {
        "A2023主映射": [], "A2023次级映射": [],
        "B2024主映射": [], "B2024次级映射": [],
    }
    for code in parsed.color_codes:
        for system in SYSTEMS:
            primary, secondary = mapping_display(index, system, code)
            if primary:
                mapping_values[f"{system}主映射"].append(f"{code}={primary}")
            if secondary:
                mapping_values[f"{system}次级映射"].append(f"{code}={secondary}")

    priority, has_sales, first_sale, last_sale, listing_count, listing_time = priority_for(
        sku, sales_stats, listing_stats, sales_status
    )
    future_auto = "否"
    if (
        not current
        and not parsed.is_bundle
        and len(parsed.color_codes) == 1
        and decision.confidence == "高"
        and decision.proposed in SYSTEMS
    ):
        future_auto = "否-低优先级" if priority.startswith("P3") else "是"

    return {
        "SKU": sku,
        "产品名称": product_name,
        "SPU": clean(source.get("spu")) or parsed.spu,
        "分类路径": clean(source.get("category_path")) or "<未分类>",
        "当前颜色体系": current,
        "SKU结构": parsed.structure + (f"（{parsed.error}）" if parsed.error else ""),
        "是否多色套装": "是" if parsed.is_bundle else "否",
        "件数": parsed.pieces,
        "版型": parsed.style,
        "颜色代码": ",".join(parsed.color_codes),
        "尺码": parsed.size,
        "品名识别颜色": "、".join(name_colors),
        "A2023主映射": "；".join(mapping_values["A2023主映射"]),
        "A2023次级映射": "；".join(mapping_values["A2023次级映射"]),
        "B2024主映射": "；".join(mapping_values["B2024主映射"]),
        "B2024次级映射": "；".join(mapping_values["B2024次级映射"]),
        "拟打颜色体系": decision.proposed or "待定",
        "置信度": decision.confidence,
        "判定类型": decision.decision_type,
        "匹配方式": decision.match_method,
        "判定原因": decision.reason,
        "历史是否有销量": has_sales,
        "首次销售日期": first_sale,
        "最近销售日期": last_sale,
        "Listing记录数": listing_count,
        "最早Listing创建时间": listing_time,
        "处理优先级": priority,
        "是否允许未来自动写入": future_auto,
    }


def mapping_quality_rows(index: MappingIndex) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for entry in index.entries:
        same_entries = index.by_system_code.get((entry.system, entry.code), [])
        other_system = "B2024" if entry.system == "A2023" else "A2023"
        other = index.by_system_code.get((other_system, entry.code), [])
        primary_chinese = same_entries[0].chinese if same_entries else ""
        issues: list[str] = []
        if not entry.chinese:
            issues.append("中文颜色为空")
        if len(same_entries) > 1:
            issues.append("同体系重复代码")
            if len({x.chinese for x in same_entries}) > 1:
                issues.append("同体系重复代码对应不同中文")
        if other:
            issues.append("A/B重叠代码")
            if primary_chinese == other[0].chinese:
                issues.append("A/B主颜色相同")
            else:
                issues.append("A/B主颜色不同")
        output.append({
            "原始位置": entry.source_position,
            "序号": entry.sequence,
            "颜色体系": entry.system,
            "颜色代码": entry.code,
            "英文": entry.english,
            "中文": entry.chinese,
            "Pantone": entry.pantone,
            "映射角色": "主映射" if entry.is_primary else "次级历史映射",
            "同体系主映射中文": primary_chinese,
            "另一体系主映射中文": other[0].chinese if other else "",
            "质量问题": "；".join(issues),
        })
    return output


def sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    priority_rank = {
        "P1-有销量": 0,
        "P2-无销量有Listing": 1,
        "P2-有Listing/销量未知": 1,
        "P3-无销量无Listing": 2,
        "P3-无Listing/销量未知": 2,
    }
    confidence_rank = {"高": 0, "中": 1, "低": 2}
    proposed_rank = {"A2023": 0, "B2024": 1, "待定": 2}
    return (
        priority_rank.get(clean(row.get("处理优先级")), 9),
        confidence_rank.get(clean(row.get("置信度")), 9),
        proposed_rank.get(clean(row.get("拟打颜色体系")), 9),
        clean(row.get("SKU")),
    )


def append_header(ws: Any, headers: Sequence[str]) -> None:
    cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cells.append(cell)
    ws.append(cells)


def append_row(ws: Any, headers: Sequence[str], row: dict[str, Any]) -> None:
    cells = []
    for header in headers:
        value = row.get(header, "")
        if isinstance(value, Decimal):
            value = float(value)
        cell = WriteOnlyCell(ws, value=value)
        if isinstance(value, (date, datetime)):
            cell.number_format = "yyyy-mm-dd"
        if header in {"判定原因", "产品名称", "SKU结构", "分类路径"}:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        cells.append(cell)
    ws.append(cells)


def configure_detail_sheet(ws: Any, row_count: int, column_count: int) -> None:
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    if row_count >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(column_count)}{row_count + 1}"
    widths = {
        "A": 24, "B": 42, "C": 16, "D": 26, "E": 14, "F": 26,
        "G": 12, "H": 10, "I": 10, "J": 22, "K": 14, "L": 28,
        "M": 28, "N": 28, "O": 28, "P": 28, "Q": 14, "R": 10,
        "S": 22, "T": 22, "U": 54, "V": 14, "W": 14, "X": 14,
        "Y": 14, "Z": 18, "AA": 22, "AB": 20,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def create_workbook(
    detail_rows: list[dict[str, Any]],
    backtest_rows: list[dict[str, Any]],
    index: MappingIndex,
    sales_status: str,
    output: Path,
) -> None:
    wb = Workbook(write_only=True)

    summary = wb.create_sheet("汇总")
    append_header(summary, ["指标", "数量/值"])

    blank_counter = Counter(row["判定类型"] for row in detail_rows)
    proposal_counter = Counter(row["拟打颜色体系"] for row in detail_rows)
    confidence_counter = Counter(row["置信度"] for row in detail_rows)
    priority_counter = Counter(row["处理优先级"] for row in detail_rows)
    auto_counter = Counter(row["是否允许未来自动写入"] for row in detail_rows)
    backtest_counter = Counter(row["回测是否命中A2023"] for row in backtest_rows)

    mapping_dup = sum(1 for values in index.by_system_code.values() if len(values) > 1)
    overlap = len(
        {code for system, code in index.by_system_code if system == "A2023"}
        & {code for system, code in index.by_system_code if system == "B2024"}
    )

    summary_rows = [
        ("安全模式", "只读dry-run；无--apply；不调用领星写接口"),
        ("颜色映射来源", COLOR_MAPPING_SOURCE),
        ("内嵌映射行数", COLOR_MAPPING_ROW_COUNT),
        ("A2023映射行数", sum(entry.system == "A2023" for entry in index.entries)),
        ("B2024映射行数", sum(entry.system == "B2024" for entry in index.entries)),
        ("同体系重复代码组", mapping_dup),
        ("A/B重叠代码数", overlap),
        ("历史销量扫描状态", sales_status),
        ("空标签分析SKU数", len(detail_rows)),
        ("已有A2023回测SKU数", len(backtest_rows)),
        ("拟打A2023", proposal_counter["A2023"]),
        ("拟打B2024", proposal_counter["B2024"]),
        ("保持待定", proposal_counter["待定"]),
        ("高置信度", confidence_counter["高"]),
        ("中置信度", confidence_counter["中"]),
        ("低置信度", confidence_counter["低"]),
        ("未来可自动写入候选", auto_counter["是"]),
        ("高置信度但P3低优先级", auto_counter["否-低优先级"]),
        ("P1有销量", priority_counter["P1-有销量"]),
        ("P2无销量有Listing", priority_counter["P2-无销量有Listing"]),
        ("P3无销量无Listing", priority_counter["P3-无销量无Listing"]),
        ("A2023回测命中", backtest_counter["是"]),
        ("A2023回测未命中", backtest_counter["否"]),
    ]
    for name, value in summary_rows:
        summary.append([name, value])
    summary.append([])
    append_header(summary, ["判定类型", "数量"])
    for name, value in blank_counter.most_common():
        summary.append([name, value])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 80
    summary.freeze_panes = "A2"

    detail = wb.create_sheet("全量判定明细")
    append_header(detail, DETAIL_HEADERS)
    for row in detail_rows:
        append_row(detail, DETAIL_HEADERS, row)
    configure_detail_sheet(detail, len(detail_rows), len(DETAIL_HEADERS))

    filters = [
        ("高置信度_A2023", lambda row: row["置信度"] == "高" and row["拟打颜色体系"] == "A2023"),
        ("高置信度_B2024", lambda row: row["置信度"] == "高" and row["拟打颜色体系"] == "B2024"),
        ("中置信度待审阅", lambda row: row["置信度"] == "中"),
        ("冲突清单", lambda row: any(key in row["判定类型"] for key in ("冲突", "同时匹配", "歧义"))),
        ("多色套装", lambda row: row["是否多色套装"] == "是"),
        ("SKU结构异常", lambda row: "异常" in row["判定类型"] or "无法解析" in row["SKU结构"]),
    ]
    for title, predicate in filters:
        subset = [row for row in detail_rows if predicate(row)]
        ws = wb.create_sheet(title)
        append_header(ws, DETAIL_HEADERS)
        for row in subset:
            append_row(ws, DETAIL_HEADERS, row)
        configure_detail_sheet(ws, len(subset), len(DETAIL_HEADERS))

    quality_headers = [
        "原始位置", "序号", "颜色体系", "颜色代码", "英文", "中文", "Pantone",
        "映射角色", "同体系主映射中文", "另一体系主映射中文", "质量问题",
    ]
    quality_rows = mapping_quality_rows(index)
    quality = wb.create_sheet("映射表质量")
    append_header(quality, quality_headers)
    for row in quality_rows:
        append_row(quality, quality_headers, row)
    quality.freeze_panes = "A2"
    quality.auto_filter.ref = f"A1:K{len(quality_rows) + 1}"
    for column, width in {
        "A": 12, "B": 10, "C": 12, "D": 12, "E": 26, "F": 26,
        "G": 24, "H": 18, "I": 26, "J": 26, "K": 48,
    }.items():
        quality.column_dimensions[column].width = width

    backtest = wb.create_sheet("A2023回测结果")
    append_header(backtest, BACKTEST_HEADERS)
    for row in backtest_rows:
        append_row(backtest, BACKTEST_HEADERS, row)
    configure_detail_sheet(backtest, len(backtest_rows), len(BACKTEST_HEADERS))
    backtest.column_dimensions["AC"].width = 18

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def run(args: argparse.Namespace) -> int:
    index = build_mapping_index()
    if len(index.entries) != COLOR_MAPPING_ROW_COUNT:
        raise RuntimeError(
            f"内嵌颜色表行数异常：expected={COLOR_MAPPING_ROW_COUNT}, actual={len(index.entries)}"
        )

    print("===== 颜色体系 SKU+品名 dry-run =====")
    print(f"内嵌颜色映射：{len(index.entries):,} 行")
    print("安全模式：只读，无 --apply，不调用领星写接口")

    db = Database()
    try:
        snapshot_rows = load_snapshot_rows(db)
        listing_stats = load_listing_stats(db)
    finally:
        db.close()

    sales_stats, sales_status = load_sales_stats(args.skip_sales)

    blank_sources: list[dict[str, Any]] = []
    a2023_sources: list[dict[str, Any]] = []
    for source in snapshot_rows:
        current = extract_color_system(source.get("custom_fields_json"))
        if not current:
            blank_sources.append(source)
        elif current == "A2023":
            a2023_sources.append(source)

    if args.limit > 0:
        blank_sources = blank_sources[: args.limit]

    print(f"产品快照SKU：{len(snapshot_rows):,}")
    print(f"颜色体系为空：{len(blank_sources):,}")
    print(f"已有A2023回测：{len(a2023_sources):,}")
    print(f"Listing可关联SKU：{len(listing_stats):,}")

    detail_rows: list[dict[str, Any]] = []
    for index_no, source in enumerate(blank_sources, start=1):
        detail_rows.append(analyze_product(source, index, sales_stats, listing_stats, sales_status))
        if index_no % 10000 == 0:
            print(f"分析进度：{index_no:,}/{len(blank_sources):,}")

    detail_rows.sort(key=sort_key)

    backtest_rows: list[dict[str, Any]] = []
    for source in a2023_sources:
        row = analyze_product(source, index, sales_stats, listing_stats, sales_status)
        row["回测是否命中A2023"] = "是" if row["拟打颜色体系"] == "A2023" else "否"
        backtest_rows.append(row)
    backtest_rows.sort(key=lambda row: (row["回测是否命中A2023"] != "否", *sort_key(row)))

    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else OUTPUT_DIR / f"color_system_name_code_analysis_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    )
    create_workbook(detail_rows, backtest_rows, index, sales_status, output)

    proposals = Counter(row["拟打颜色体系"] for row in detail_rows)
    confidence = Counter(row["置信度"] for row in detail_rows)
    priority = Counter(row["处理优先级"] for row in detail_rows)
    auto = Counter(row["是否允许未来自动写入"] for row in detail_rows)
    backtest = Counter(row["回测是否命中A2023"] for row in backtest_rows)

    print("===== dry-run汇总 =====")
    print(f"拟打A2023：{proposals['A2023']:,}")
    print(f"拟打B2024：{proposals['B2024']:,}")
    print(f"保持待定：{proposals['待定']:,}")
    print(f"高/中/低置信度：{confidence['高']:,}/{confidence['中']:,}/{confidence['低']:,}")
    print(f"P1有销量：{priority['P1-有销量']:,}")
    print(f"P2无销量有Listing：{priority['P2-无销量有Listing']:,}")
    print(f"P3无销量无Listing：{priority['P3-无销量无Listing']:,}")
    print(f"未来可自动写入候选：{auto['是']:,}")
    print(f"高置信度但P3低优先级：{auto['否-低优先级']:,}")
    print(f"A2023回测命中/未命中：{backtest['是']:,}/{backtest['否']:,}")
    print(f"输出：{output}")

    show = max(0, int(args.show))
    if show:
        print("\n前几条高置信度候选：")
        shown = 0
        for row in detail_rows:
            if row["置信度"] != "高":
                continue
            print(
                f"{row['SKU']}\t{row['颜色代码']}\t{row['品名识别颜色']}\t"
                f"{row['拟打颜色体系']}\t{row['处理优先级']}"
            )
            shown += 1
            if shown >= show:
                break
    return 0


def main(argv: Sequence[str] | None = None) -> int:
    return run(parse_args(argv))


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n用户中断")
        sys.exit(130)
    except Exception as exc:
        print(f"执行失败：{type(exc).__name__}: {exc}")
        raise
