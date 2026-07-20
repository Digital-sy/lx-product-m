#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系高置信度第一轮：生成审阅清单并安全写入 A2023。

安全边界：
1. prepare 只从 V3 dry-run Excel 选择高置信度 A2023，默认 100 条。
2. 只选择 P1/P2、单颜色、主映射精确唯一匹配、非特殊后缀 SKU。
3. 默认尽量每个 SPU 只选一条，扩大首轮验证覆盖面。
4. apply 写前实时读取领星；颜色体系非空且不同于 A2023 时拒绝覆盖。
5. 完整保留品名、分类和其他自定义字段，写后再次实时复查。
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from lx_product_m.db import Database
from lx_product_m.lingxing_client import LingxingClient
from lx_product_m.services.product_write_guard import (
    PRODUCT_SET_API,
    build_guarded_product_set_body,
    clean,
    extract_custom_fields,
    fetch_live_products,
    request_with_retry,
    target_fields_match,
)
from scripts.color_system_tagging import (
    BEIJING_TZ,
    COLOR_FIELD_ID_ENV,
    COLOR_FIELD_NAME,
    OUTPUT_DIR,
    RATE_LIMIT_RETRIES,
    RATE_LIMIT_WAIT_SECONDS,
    create_failure_workbook,
    discover_field_id_from_live,
    failure_record,
    field_id_from_snapshot,
    field_ids_from_fields,
    field_value,
    is_beijing_low_peak,
    load_review_file,
    print_apply_progress,
    response_succeeded,
    validate_field_id,
    verify_product_set_result,
)

TARGET_VALUE = "A2023"
DEFAULT_FIELD_ID = "207722905719915521"
REVIEW_SHEET = "拟打标明细"
SOURCE_SHEET = "全量判定明细"

REQUIRED_SOURCE_HEADERS = {
    "SKU",
    "产品名称",
    "SPU",
    "当前颜色体系",
    "SKU结构",
    "是否多色套装",
    "颜色代码",
    "品名识别颜色",
    "拟打颜色体系",
    "置信度",
    "判定类型",
    "匹配方式",
    "处理优先级",
    "是否允许未来自动写入",
}

REVIEW_HEADERS = [
    "SKU",
    "拟打值",
    "产品名称",
    "SPU",
    "颜色代码",
    "品名识别颜色",
    "处理优先级",
    "判定类型",
    "匹配方式",
    "SKU结构",
    "来源分析文件",
]


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="颜色体系高置信度 A2023 第一轮")
    sub = parser.add_subparsers(dest="command", required=True)

    prepare = sub.add_parser("prepare", help="从 V3 dry-run 生成首轮审阅清单")
    prepare.add_argument("--analysis-file", required=True, help="V3 dry-run Excel")
    prepare.add_argument("--output", default="", help="审阅清单输出路径")
    prepare.add_argument("--limit", type=int, default=100, help="首轮数量，默认100")
    prepare.add_argument(
        "--allow-multiple-name-colors",
        action="store_true",
        help="允许品名识别到多个颜色；首轮默认排除",
    )

    apply = sub.add_parser("apply", help="写入人工确认后的审阅清单")
    apply.add_argument("--review-file", required=True)
    apply.add_argument("--expected-count", type=int, required=True, help="审阅清单预期条数，不一致则终止")
    apply.add_argument("--field-id", default=DEFAULT_FIELD_ID)
    apply.add_argument("--batch-size", type=int, default=100)
    apply.add_argument("--delay", type=float, default=0.5)
    apply.add_argument("--verify-delay", type=float, default=1.0)
    apply.add_argument("--allow-outside-low-peak", action="store_true")
    return parser.parse_args(argv)


def _headers(sheet: Any) -> dict[str, int]:
    first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return {clean(value): idx for idx, value in enumerate(first) if clean(value)}


def _cell(cells: Sequence[Any], headers: dict[str, int], name: str) -> str:
    idx = headers[name]
    return clean(cells[idx] if idx < len(cells) else "")


def eligibility_reason(row: dict[str, str], *, allow_multiple_name_colors: bool = False) -> str:
    if row["当前颜色体系"]:
        return "当前颜色体系非空"
    if row["拟打颜色体系"] != TARGET_VALUE:
        return "不是A2023"
    if row["置信度"] != "高":
        return "不是高置信度"
    if row["判定类型"] != "精确唯一匹配":
        return "不是主映射精确唯一匹配"
    if row["是否多色套装"] == "是":
        return "多色套装"
    if row["是否允许未来自动写入"] != "是":
        return "分析结果未允许自动写入"
    if not (row["处理优先级"].startswith("P1-") or row["处理优先级"].startswith("P2-")):
        return "P3低优先级"
    structure = row["SKU结构"]
    if any(token in structure for token in ("特殊后缀", "未收录", "异常", "无法解析", "多颜色")):
        return "特殊或异常SKU结构"
    if not row["颜色代码"] or "," in row["颜色代码"]:
        return "颜色代码不唯一"
    name_colors = row["品名识别颜色"]
    if not name_colors:
        return "品名未识别颜色"
    if not allow_multiple_name_colors and any(separator in name_colors for separator in ("、", ",", "，", ";", "；")):
        return "品名识别多个颜色"
    return ""


def load_source_rows(path: Path) -> tuple[list[dict[str, str]], Counter[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in wb.sheetnames:
            raise ValueError(f"分析文件缺少 sheet：{SOURCE_SHEET}")
        sheet = wb[SOURCE_SHEET]
        headers = _headers(sheet)
        missing = sorted(REQUIRED_SOURCE_HEADERS - set(headers))
        if missing:
            raise ValueError("分析文件缺少列：" + ",".join(missing))

        rows: list[dict[str, str]] = []
        duplicates: Counter[str] = Counter()
        for cells in sheet.iter_rows(min_row=2, values_only=True):
            sku = _cell(cells, headers, "SKU")
            if not sku:
                continue
            duplicates[sku] += 1
            rows.append({name: _cell(cells, headers, name) for name in REQUIRED_SOURCE_HEADERS})

        repeated = [sku for sku, count in duplicates.items() if count > 1]
        if repeated:
            raise ValueError(f"分析文件存在重复SKU，共{len(repeated)}个，例如：{repeated[:5]}")
        return rows, duplicates
    finally:
        wb.close()


def select_review_rows(
    rows: Sequence[dict[str, str]],
    *,
    limit: int,
    allow_multiple_name_colors: bool = False,
) -> tuple[list[dict[str, str]], Counter[str], int]:
    if limit <= 0:
        raise ValueError("--limit 必须大于0")

    exclusions: Counter[str] = Counter()
    eligible: list[dict[str, str]] = []
    for row in rows:
        reason = eligibility_reason(row, allow_multiple_name_colors=allow_multiple_name_colors)
        if reason:
            exclusions[reason] += 1
        else:
            eligible.append(dict(row))

    def sort_key(row: dict[str, str]) -> tuple[Any, ...]:
        priority = 0 if row["处理优先级"].startswith("P1-") else 1
        return priority, row["颜色代码"], row["SPU"], row["SKU"]

    eligible.sort(key=sort_key)

    # 首轮优先覆盖不同 SPU，避免一个款的多个尺码占满100条。
    selected: list[dict[str, str]] = []
    selected_skus: set[str] = set()
    seen_spu: set[str] = set()
    for row in eligible:
        spu = row["SPU"] or row["SKU"].split("-", 1)[0]
        if spu in seen_spu:
            continue
        selected.append(row)
        selected_skus.add(row["SKU"])
        seen_spu.add(spu)
        if len(selected) >= limit:
            break

    if len(selected) < limit:
        for row in eligible:
            if row["SKU"] in selected_skus:
                continue
            selected.append(row)
            selected_skus.add(row["SKU"])
            if len(selected) >= limit:
                break

    return selected, exclusions, len(eligible)


def create_review_workbook(
    selected: Sequence[dict[str, str]],
    exclusions: Counter[str],
    eligible_count: int,
    source_path: Path,
    output: Path,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = REVIEW_SHEET
    sheet.append(REVIEW_HEADERS)
    for row in selected:
        values = {
            **row,
            "拟打值": TARGET_VALUE,
            "来源分析文件": source_path.name,
        }
        sheet.append([values.get(header, "") for header in REVIEW_HEADERS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    summary = wb.create_sheet("汇总")
    summary.append(["项目", "数量/内容"])
    summary.append(["来源分析文件", str(source_path)])
    summary.append(["符合全部安全条件", eligible_count])
    summary.append(["本轮选中", len(selected)])
    summary.append(["本轮目标值", TARGET_VALUE])
    summary.append(["选择策略", "P1优先、P2其次、优先不同SPU"])
    summary.append([])
    summary.append(["排除原因", "数量"])
    for reason, count in exclusions.most_common():
        summary.append([reason, count])

    for ws in wb.worksheets:
        for column in ws.columns:
            width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            ws.column_dimensions[column[0].column_letter].width = width
    wb.save(output)


def prepare(args: argparse.Namespace) -> int:
    analysis_path = Path(args.analysis_file).expanduser().resolve()
    if not analysis_path.exists():
        raise FileNotFoundError(f"V3分析文件不存在：{analysis_path}")
    rows, _ = load_source_rows(analysis_path)
    selected, exclusions, eligible_count = select_review_rows(
        rows,
        limit=int(args.limit),
        allow_multiple_name_colors=bool(args.allow_multiple_name_colors),
    )
    if len(selected) != int(args.limit):
        raise RuntimeError(
            f"符合条件的SKU不足：计划{int(args.limit):,}，实际选中{len(selected):,}，"
            f"符合全部条件共{eligible_count:,}"
        )

    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else OUTPUT_DIR / f"color_system_high_confidence_A2023_round1_{len(selected)}_{datetime.now(BEIJING_TZ):%Y%m%d_%H%M%S}.xlsx"
    )
    create_review_workbook(selected, exclusions, eligible_count, analysis_path, output)
    print("===== 高置信度A2023第一轮清单 =====")
    print(f"来源：{analysis_path}")
    print(f"符合安全条件：{eligible_count:,}")
    print(f"本轮选中：{len(selected):,}")
    print(f"不同SPU：{len({row['SPU'] or row['SKU'].split('-', 1)[0] for row in selected}):,}")
    print(f"P1：{sum(row['处理优先级'].startswith('P1-') for row in selected):,}")
    print(f"P2：{sum(row['处理优先级'].startswith('P2-') for row in selected):,}")
    print(f"输出：{output}")
    return 0


def live_color_values(fields: Any, field_id: str) -> set[str]:
    values: set[str] = set()
    for item in fields if isinstance(fields, list) else []:
        if not isinstance(item, dict):
            continue
        item_id = clean(item.get("id"))
        name = clean(item.get("name"))
        if item_id != field_id and name != COLOR_FIELD_NAME:
            continue
        value = field_value(item)
        if value:
            values.add(value)
    return values


async def apply_review(args: argparse.Namespace) -> int:
    if not args.allow_outside_low_peak and not is_beijing_low_peak():
        now = datetime.now(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")
        raise RuntimeError(
            f"当前北京时间 {now} 不在低峰窗口00:00-06:00；"
            "请在凌晨执行，或人工确认后使用 --allow-outside-low-peak"
        )

    review_path = Path(args.review_file).expanduser().resolve()
    targets = load_review_file(review_path)
    if len(targets) != int(args.expected_count):
        raise ValueError(
            f"审阅清单数量不一致：expected={int(args.expected_count):,}, actual={len(targets):,}"
        )
    illegal = [item for item in targets if item["拟打值"] != TARGET_VALUE]
    if illegal:
        raise ValueError(f"第一轮只允许写入 {TARGET_VALUE}，发现非法行：{illegal[:3]}")

    batch_size = max(1, min(100, int(args.batch_size)))
    field_id = validate_field_id(args.field_id or os.getenv(COLOR_FIELD_ID_ENV, ""))
    failures: list[dict[str, str]] = []
    stats: Counter[str] = Counter(processed=len(targets))

    print("===== 高置信度A2023第一轮安全写入 =====")
    print(f"审阅清单：{review_path}")
    print(f"计划处理：{len(targets):,}")
    print("写前保护：实时颜色体系为空才写；相同跳过；不同值拒绝覆盖")
    print("写入保护：完整字段合并 + product/set + 写后实时复查")

    db = Database()
    client = LingxingClient(db=db, enable_api_log=True)
    try:
        token = (await client.generate_token()).token
        if not field_id:
            field_id = field_id_from_snapshot(db)
        if not field_id:
            field_id, token = await discover_field_id_from_live(
                client, token, [item["SKU"] for item in targets]
            )
        if not field_id:
            raise RuntimeError("无法发现颜色体系字段ID")
        print(f"颜色体系字段 ID：{field_id}")

        verify_delay = max(0.0, float(args.verify_delay))
        for start in range(0, len(targets), batch_size):
            batch = targets[start : start + batch_size]
            try:
                live_map, token = await fetch_live_products(
                    client,
                    token,
                    [item["SKU"] for item in batch],
                    max_retries=RATE_LIMIT_RETRIES,
                    batch_size=batch_size,
                    delay_seconds=0.5,
                    retry_base_seconds=RATE_LIMIT_WAIT_SECONDS,
                    retry_max_seconds=RATE_LIMIT_WAIT_SECONDS,
                )
            except Exception as exc:
                for item in batch:
                    failures.append(failure_record(item["SKU"], TARGET_VALUE, "写前实时读取", exc))
                    stats["failed"] += 1
                print_apply_progress(stats, len(targets))
                continue

            pending_verify: list[tuple[str, str, dict[str, Any]]] = []
            for item in batch:
                sku = item["SKU"]
                target = {"id": field_id, "name": COLOR_FIELD_NAME, "val": TARGET_VALUE}
                try:
                    product = live_map.get(sku)
                    if not product:
                        raise RuntimeError("写前实时详情未返回该SKU")
                    fields = extract_custom_fields(product)
                    discovered = field_ids_from_fields(fields)
                    if discovered and discovered != {field_id}:
                        raise RuntimeError(
                            f"实时详情字段ID与配置不一致：configured={field_id}, live={sorted(discovered)}"
                        )

                    current_values = live_color_values(fields, field_id)
                    if current_values == {TARGET_VALUE}:
                        stats["skipped"] += 1
                        continue
                    if current_values:
                        message = "实时颜色体系已存在不同值，拒绝覆盖：" + ",".join(sorted(current_values))
                        failures.append(failure_record(sku, TARGET_VALUE, "写前冲突保护", message))
                        stats["failed"] += 1
                        continue
                    if target_fields_match(fields, [target]):
                        stats["skipped"] += 1
                        continue

                    body, _ = build_guarded_product_set_body(
                        product,
                        sku=sku,
                        target_custom_fields=[target],
                        preserve_current_category=True,
                    )
                    response, token = await request_with_retry(
                        client,
                        token,
                        PRODUCT_SET_API,
                        body,
                        max_retries=RATE_LIMIT_RETRIES,
                        retry_base_seconds=RATE_LIMIT_WAIT_SECONDS,
                        retry_max_seconds=RATE_LIMIT_WAIT_SECONDS,
                    )
                    if response_succeeded(response):
                        pending_verify.append((sku, TARGET_VALUE, body))
                    else:
                        failures.append(
                            failure_record(sku, TARGET_VALUE, "product/set", response, response=response)
                        )
                        stats["failed"] += 1
                except Exception as exc:
                    failures.append(failure_record(sku, TARGET_VALUE, "安全合并写入", exc))
                    stats["failed"] += 1
                if args.delay > 0:
                    await asyncio.sleep(args.delay)

            if pending_verify:
                if verify_delay > 0:
                    await asyncio.sleep(verify_delay)
                try:
                    verify_map, token = await fetch_live_products(
                        client,
                        token,
                        [item[0] for item in pending_verify],
                        max_retries=RATE_LIMIT_RETRIES,
                        batch_size=batch_size,
                        delay_seconds=0.5,
                        retry_base_seconds=RATE_LIMIT_WAIT_SECONDS,
                        retry_max_seconds=RATE_LIMIT_WAIT_SECONDS,
                    )
                except Exception as exc:
                    for sku, value, _ in pending_verify:
                        failures.append(failure_record(sku, value, "写后实时复查", exc))
                        stats["failed"] += 1
                else:
                    for sku, value, body in pending_verify:
                        ok, message = verify_product_set_result(verify_map.get(sku), body)
                        if ok:
                            stats["success"] += 1
                        else:
                            failures.append(failure_record(sku, value, "写后复查", message))
                            stats["failed"] += 1
            print_apply_progress(stats, len(targets))
    finally:
        await client.aclose()
        db.close()

    failure_path: Path | None = None
    if failures:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        failure_path = OUTPUT_DIR / f"color_system_high_confidence_round1_failures_{datetime.now(BEIJING_TZ):%Y%m%d_%H%M%S}.xlsx"
        create_failure_workbook(failures, failure_path)

    print(
        f"执行统计：处理={stats['processed']:,} 成功={stats['success']:,} "
        f"失败={stats['failed']:,} 跳过={stats['skipped']:,}"
    )
    if failure_path:
        print(f"失败清单：{failure_path}")
    return 1 if failures else 0


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if args.command == "prepare":
        return prepare(args)
    return asyncio.run(apply_review(args))


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n用户中断")
        sys.exit(130)
    except Exception as exc:
        print(f"执行失败：{type(exc).__name__}: {exc}")
        raise
