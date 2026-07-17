#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""颜色体系批量打标（第二轮：待定组按开发年份消化）。

默认只生成审阅清单。范围必须来自第一轮 Excel 中拟打值为“待定”的 SKU；
正式写入只提交审阅后拟打值为 A2023 的行，其余行保持不动。

开发年份判定规则：原值为“历史”，或清洗后可解析为整数年份且 <= 2024，
转 A2023；清洗后可解析为整数年份且 >= 2025、空值、无法解析的意外值均
维持待定，并继续输出意外值清单。清洗包含 strip 空白、去除“年”字后缀、
全角转半角。
"""
from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from lx_product_m.db import Database
from lx_product_m.services.product_write_guard import clean
from scripts.color_system_tagging import (
    BEIJING_TZ,
    OUTPUT_DIR,
    apply_review_file,
    auto_width,
    load_review_file,
    style_header,
)

DEVELOPMENT_YEAR_FIELD_NAME = "开发年份"
DEVELOPMENT_YEAR_FIELD_ID = "207714670595318273"
EXPECTED_PENDING_COUNT = 9823
DETAIL_HEADERS = ["SKU", "开发年份原值", "拟打值"]

STATUS_CONVERT = "转 A2023"
STATUS_KEEP_FUTURE = "维持待定（2025+）"
STATUS_KEEP_EMPTY = "维持待定（空值）"
STATUS_KEEP_UNEXPECTED = "维持待定（意外值）"

SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FONT = Font(name="Microsoft YaHei", bold=True)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="颜色体系批量打标（第二轮：待定组按开发年份消化）")
    parser.add_argument(
        "--first-round-file",
        default="",
        help="dry-run 必填：第一轮人工确认范围 Excel",
    )
    parser.add_argument(
        "--expected-pending-count",
        type=int,
        default=EXPECTED_PENDING_COUNT,
        help="第一轮待定 SKU 预期数量；默认9823，设0可关闭数量断言",
    )
    parser.add_argument("--output", default="", help="第二轮 dry-run Excel 输出路径")
    parser.add_argument("--db-batch-size", type=int, default=500, help="读取产品快照的 SKU 批次")
    parser.add_argument("--show", type=int, default=30, help="控制台明细预览行数")
    parser.add_argument("--apply", action="store_true", help="写入人工审阅后的第二轮清单")
    parser.add_argument("--review-file", default="", help="--apply 时必填：第二轮审阅 Excel")
    parser.add_argument("--field-id", default="", help="领星“颜色体系”字段 ID")
    parser.add_argument("--batch-size", type=int, default=100, help="写前实时读取批次，最大100")
    parser.add_argument("--delay", type=float, default=0.5, help="每个实际写入 SKU 后等待秒数")
    parser.add_argument(
        "--allow-outside-low-peak",
        action="store_true",
        help="允许在北京时间 00:00-06:00 之外执行 --apply",
    )
    return parser.parse_args(argv)


def load_first_round_pending_skus(path: Path, expected_count: int = EXPECTED_PENDING_COUNT) -> list[str]:
    rows = load_review_file(path)
    skus = [row["SKU"] for row in rows if row["拟打值"] == "待定"]
    if expected_count < 0:
        raise ValueError("--expected-pending-count 不能小于0")
    if expected_count and len(skus) != expected_count:
        raise RuntimeError(
            f"第一轮待定 SKU 数量不符：expected={expected_count:,}, actual={len(skus):,}；"
            "请确认输入的是完整且正确的第一轮审阅文件"
        )
    return skus


def parse_custom_fields(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError:
            return []
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def original_field_value(item: dict[str, Any]) -> str:
    saw_empty = False
    for key in ("val_text", "val", "value", "field_value"):
        if key in item and item.get(key) is not None:
            value = str(item.get(key))
            if value != "":
                return value
            saw_empty = True
    if saw_empty:
        return ""
    return ""


def extract_development_year(value: Any) -> str:
    """保留原始值，不 strip；脏空格必须作为意外值暴露。"""
    values: list[str] = []
    for item in parse_custom_fields(value):
        if (
            clean(item.get("name")) == DEVELOPMENT_YEAR_FIELD_NAME
            or clean(item.get("id")) == DEVELOPMENT_YEAR_FIELD_ID
        ):
            raw = original_field_value(item)
            if raw not in values:
                values.append(raw)
    if not values:
        return ""
    if len(values) == 1:
        return values[0]
    return "[多值]" + "|".join(values)


def clean_development_year(raw_value: Any) -> str:
    value = "" if raw_value is None else str(raw_value)
    value = unicodedata.normalize("NFKC", value).strip()
    if value.endswith("年"):
        value = value[:-1].strip()
    return value


def classify_development_year(raw_value: Any) -> tuple[str, str]:
    value = clean_development_year(raw_value)
    if value == "历史":
        return "A2023", STATUS_CONVERT
    if value == "":
        return "待定", STATUS_KEEP_EMPTY
    if re.fullmatch(r"\d+", value):
        year = int(value)
        if year <= 2024:
            return "A2023", STATUS_CONVERT
        return "待定", STATUS_KEEP_FUTURE
    return "待定", STATUS_KEEP_UNEXPECTED


def load_development_year_values(
    db: Database,
    skus: Sequence[str],
    batch_size: int = 500,
) -> dict[str, str]:
    batch_size = max(1, min(int(batch_size), 1000))
    values = {sku: "" for sku in skus}
    for start in range(0, len(skus), batch_size):
        batch = list(skus[start : start + batch_size])
        placeholders = ",".join(["%s"] * len(batch))
        rows = db.fetch_all(
            f"""
            SELECT sku, custom_fields_json
            FROM lxpm_product_category_snapshot
            WHERE sku IN ({placeholders})
            """,
            batch,
        )
        for row in rows:
            sku = clean(row.get("sku"))
            if sku in values:
                values[sku] = extract_development_year(row.get("custom_fields_json"))
        print(f"开发年份读取进度：{min(start + batch_size, len(skus)):,}/{len(skus):,}")
    return values


def group_actual_values(skus: Iterable[str], values: dict[str, str]) -> Counter[str]:
    return Counter(values.get(sku, "") for sku in skus)


def prepare_round2_rows(
    skus: Iterable[str],
    values: dict[str, str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for sku in skus:
        raw = values.get(sku, "")
        proposed, status = classify_development_year(raw)
        rows.append(
            {
                "SKU": sku,
                "开发年份原值": raw,
                "拟打值": proposed,
                "_判定": status,
            }
        )
    rows.sort(
        key=lambda row: (
            0 if row["拟打值"] == "A2023" else 1,
            row["开发年份原值"],
            row["SKU"],
        )
    )
    return rows


def append_section_title(sheet: Any, title: str) -> int:
    row_no = sheet.max_row + 1
    sheet.append([title])
    sheet.cell(row_no, 1).fill = SECTION_FILL
    sheet.cell(row_no, 1).font = SECTION_FONT
    return row_no


def raw_value_label(value: str) -> str:
    return value if value != "" else "<空值>"


def create_round2_workbook(rows: Sequence[dict[str, str]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    detail = workbook.active
    detail.title = "拟打标明细"
    detail.append(DETAIL_HEADERS)
    for row in rows:
        detail.append([row.get(header, "") for header in DETAIL_HEADERS])
    style_header(detail)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    if detail.max_row >= 2:
        validation = DataValidation(type="list", formula1='"A2023,待定"', allow_blank=False)
        validation.error = "拟打值只允许 A2023 或 待定"
        validation.errorTitle = "非法拟打值"
        detail.add_data_validation(validation)
        validation.add(f"C2:C{detail.max_row}")
    auto_width(detail)

    summary = workbook.create_sheet("汇总")
    summary.append(["项目", "SKU数"])
    proposed_counts = Counter(row["拟打值"] for row in rows)
    status_counts = Counter(row["_判定"] for row in rows)
    summary.append(["转 A2023", proposed_counts["A2023"]])
    summary.append(["剩余待定", proposed_counts["待定"]])
    summary.append(["空值", status_counts[STATUS_KEEP_EMPTY]])
    summary.append(["意外值", status_counts[STATUS_KEEP_UNEXPECTED]])
    style_header(summary, 1)

    actual_counts = Counter(row["开发年份原值"] for row in rows)
    statuses_by_value: dict[str, str] = {}
    for row in rows:
        statuses_by_value[row["开发年份原值"]] = row["_判定"]

    summary.append([])
    append_section_title(summary, "开发年份实际值分布（原值未清洗）")
    distribution_header = summary.max_row + 1
    summary.append(["开发年份原值", "SKU数", "规则结果"])
    for raw_value in sorted(actual_counts, key=lambda value: (value != "", value)):
        summary.append(
            [raw_value_label(raw_value), actual_counts[raw_value], statuses_by_value[raw_value]]
        )
    style_header(summary, distribution_header)

    unexpected_skus: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        if row["_判定"] == STATUS_KEEP_UNEXPECTED:
            unexpected_skus[row["开发年份原值"]].append(row["SKU"])

    summary.append([])
    append_section_title(summary, "意外值清单")
    unexpected_header = summary.max_row + 1
    summary.append(["开发年份原值", "SKU数", "SKU示例（最多20个）"])
    if unexpected_skus:
        for raw_value in sorted(unexpected_skus):
            sku_list = unexpected_skus[raw_value]
            summary.append(
                [raw_value_label(raw_value), len(sku_list), ",".join(sku_list[:20])]
            )
    else:
        summary.append(["（无）", 0, ""])
    style_header(summary, unexpected_header)
    summary.freeze_panes = "A2"
    auto_width(summary, max_width=80)
    workbook.save(output)


def print_actual_value_distribution(counts: Counter[str]) -> None:
    print("===== 开发年份实际值 GROUP BY（原值未清洗） =====")
    for raw_value in sorted(counts, key=lambda value: (value != "", value)):
        print(f"{raw_value!r}: {counts[raw_value]:,}")


def run_dry_run(args: argparse.Namespace) -> int:
    if not args.first_round_file:
        raise ValueError("第二轮 dry-run 必须指定 --first-round-file 以锁定第一轮待定范围")
    if args.review_file:
        raise ValueError("--review-file 只允许与 --apply 一起使用")

    first_round_path = Path(args.first_round_file).expanduser().resolve()
    skus = load_first_round_pending_skus(first_round_path, args.expected_pending_count)
    print("===== 颜色体系批量打标 dry-run（第二轮） =====")
    print(f"第一轮范围文件：{first_round_path}")
    print(f"第一轮待定 SKU：{len(skus):,}")

    db = Database()
    try:
        values = load_development_year_values(db, skus, args.db_batch_size)
    finally:
        db.close()

    actual_counts = group_actual_values(skus, values)
    print_actual_value_distribution(actual_counts)
    rows = prepare_round2_rows(skus, values)
    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else OUTPUT_DIR
        / f"color_system_tagging_round2_dryrun_{datetime.now(BEIJING_TZ):%Y%m%d}.xlsx"
    )
    create_round2_workbook(rows, output)

    for row in rows[: max(0, args.show)]:
        print(f"{row['SKU']} | {row['开发年份原值']!r} -> {row['拟打值']}")
    counts = Counter(row["拟打值"] for row in rows)
    unexpected_count = sum(row["_判定"] == STATUS_KEEP_UNEXPECTED for row in rows)
    print(
        "执行统计："
        f"处理={len(rows):,} 转A2023={counts['A2023']:,} "
        f"剩余待定={counts['待定']:,} 意外值={unexpected_count:,}"
    )
    print(f"第二轮 dry-run 清单：{output}")
    return 0


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if args.apply:
        if args.first_round_file:
            raise ValueError("--apply 只读取 --review-file，不应再指定 --first-round-file")
        return asyncio.run(
            apply_review_file(
                args,
                write_values={"A2023"},
                failure_prefix="color_system_tagging_round2_failures",
            )
        )
    return run_dry_run(args)


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n用户中断")
        sys.exit(130)
    except Exception as exc:  # noqa: BLE001
        print(f"执行失败：{type(exc).__name__}: {exc}")
        raise
