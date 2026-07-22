from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts import color_system_name_code_analysis as analysis


class MappingTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.index = analysis.build_mapping_index()

    def test_embedded_mapping_counts_and_first_row_priority(self) -> None:
        self.assertEqual(len(self.index.entries), 393)
        self.assertEqual(
            sum(entry.system == "A2023" for entry in self.index.entries),
            194,
        )
        self.assertEqual(
            sum(entry.system == "B2024" for entry in self.index.entries),
            199,
        )
        a_sn = self.index.by_system_code[("A2023", "SN")]
        self.assertEqual(a_sn[0].chinese, "浅赭色")
        self.assertTrue(a_sn[0].is_primary)
        self.assertEqual(a_sn[1].chinese, "赭色")
        self.assertFalse(a_sn[1].is_primary)

    def test_overlapping_code_can_be_distinguished_by_product_name(self) -> None:
        a = analysis.decide_single_code("BO", "测试亮橘色上衣", self.index)
        b = analysis.decide_single_code("BO", "测试黑玛瑙上衣", self.index)
        self.assertEqual((a.proposed, a.confidence), ("A2023", "高"))
        self.assertEqual((b.proposed, b.confidence), ("B2024", "高"))

    def test_same_code_same_color_across_systems_stays_pending(self) -> None:
        result = analysis.decide_single_code("BAB", "低音蓝连衣裙", self.index)
        self.assertEqual(result.proposed, "")
        self.assertEqual(result.decision_type, "A/B同时匹配")

    def test_secondary_mapping_is_medium_confidence(self) -> None:
        result = analysis.decide_single_code(
            "SN",
            "NY006-凸点钢圈内衣赭色-40DD码",
            self.index,
        )
        self.assertEqual(result.proposed, "A2023")
        self.assertEqual(result.confidence, "中")
        self.assertEqual(result.decision_type, "次级映射匹配")

    def test_missing_color_suffix_is_supported(self) -> None:
        result = analysis.decide_single_code("BO", "亮橘短款上衣", self.index)
        self.assertEqual(result.proposed, "A2023")
        self.assertEqual(result.confidence, "高")


class SkuParsingTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.index = analysis.build_mapping_index()

    def test_regular_short_glued_and_bundle(self) -> None:
        regular = analysis.parse_sku("NY006-SN-40DD", self.index)
        self.assertEqual(regular.color_codes, ("SN",))
        self.assertEqual(regular.size, "40DD")

        short = analysis.parse_sku("KZ291-SHORT-BK-S", self.index)
        self.assertEqual(short.style, "SHORT")
        self.assertEqual(short.color_codes, ("BK",))
        self.assertEqual(short.size, "S")

        glued = analysis.parse_sku("ABC-SN40DD", self.index)
        self.assertEqual(glued.structure, "颜色尺码粘连")
        self.assertEqual(glued.color_codes, ("SN",))
        self.assertEqual(glued.size, "40DD")

        bundle = analysis.parse_sku(
            "NKW033-5PCS-WH-BE-SD-GY-MGR-XL",
            self.index,
        )
        self.assertTrue(bundle.is_bundle)
        self.assertEqual(bundle.pieces, "5PCS")
        self.assertEqual(bundle.color_codes, ("WH", "BE", "SD", "GY", "MGR"))
        self.assertEqual(bundle.size, "XL")

    def test_unknown_code_is_structure_error(self) -> None:
        result = analysis.parse_sku("ABC-ZZZ-XL", self.index)
        self.assertTrue(result.error)
        self.assertEqual(result.color_codes, ())


class SafetyAndWorkbookTest(unittest.TestCase):
    def test_apply_argument_is_not_available(self) -> None:
        with self.assertRaises(SystemExit):
            analysis.parse_args(["--apply"])

    def test_small_workbook_contains_expected_sheets(self) -> None:
        index = analysis.build_mapping_index()
        source = {
            "sku": "TEST-BO-M",
            "product_name": "测试亮橘色上衣",
            "spu": "TEST",
            "category_path": "24/春夏/特征款",
            "custom_fields_json": "[]",
        }
        row = analysis.analyze_product(
            source,
            index,
            sales_stats={"TEST-BO-M": {"has_sales": True}},
            listing_stats={
                "TEST-BO-M": {
                    "listing_count": 1,
                    "earliest_listing_create_time": None,
                }
            },
            sales_status="成功：2024",
        )
        backtest = dict(row)
        backtest["当前颜色体系"] = "A2023"
        backtest["回测是否命中A2023"] = "是"

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "dryrun.xlsx"
            analysis.create_workbook([row], [backtest], index, "成功：2024", output)
            wb = load_workbook(output, read_only=True, data_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "汇总",
                        "全量判定明细",
                        "高置信度_A2023",
                        "高置信度_B2024",
                        "中置信度待审阅",
                        "冲突清单",
                        "多色套装",
                        "SKU结构异常",
                        "映射表质量",
                        "A2023回测结果",
                    ],
                )
                self.assertEqual(
                    wb["高置信度_A2023"]["A2"].value,
                    "TEST-BO-M",
                )
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
