from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

from openpyxl import load_workbook

from lx_product_m.services import product_write_guard
from scripts import color_system_tagging as tagging


class ColorSystemRulesTest(unittest.TestCase):
    def test_first_round_boundary(self) -> None:
        self.assertEqual(tagging.classify_opening_date(date(2024, 6, 30)), "A2023")
        self.assertIsNone(tagging.classify_opening_date(date(2024, 7, 1)))
        self.assertEqual(tagging.classify_opening_date(None), "待定")

    def test_prepare_rows_keeps_targets_and_counts_later_sales_as_skipped(self) -> None:
        rows = [
            {
                "sku": "AA100-BK-S",
                "associated_msku": "AA100-BK-S,AA100-BK-S-FBA",
                "opening_date": date(2024, 6, 30),
                "first_sale_stores": "sy_us",
                "has_recent_sales": 1,
            },
            {
                "sku": "BB200-WH-M",
                "associated_msku": "BB200-WH-M",
                "opening_date": None,
                "first_sale_stores": None,
                "has_recent_sales": 0,
            },
            {
                "sku": "CC300-RD-L",
                "associated_msku": "CC300-RD-L",
                "opening_date": date(2024, 7, 1),
                "first_sale_stores": "jq_us",
                "has_recent_sales": 1,
            },
        ]

        prepared, skipped = tagging.prepare_analysis_rows(rows)

        self.assertEqual(skipped, 1)
        self.assertEqual([row["拟打值"] for row in prepared], ["A2023", "待定"])
        self.assertEqual(prepared[0]["SPU"], "AA100")
        self.assertEqual(prepared[0]["近半年是否有销量"], "是")
        self.assertEqual(prepared[1]["近半年是否有销量"], "否")

    def test_sql_reuses_order_filters_and_groups_first_sale_by_msku(self) -> None:
        sql = tagging.color_system_sql(["sy_order.total_order_2024"], ["sy_us"])

        self.assertIn("CAST(TRIM(CAST(o.type AS CHAR)) AS BINARY) = 0x4F72646572", sql)
        self.assertIn("COALESCE(CAST(o.quantity AS DECIMAL(18,4)), 0) > 0", sql)
        self.assertIn("COALESCE(CAST(o.`product sales` AS DECIMAL(18,4)), 0) > 0", sql)
        self.assertIn("<> 0x533031", sql)
        self.assertIn("WHEN LEFT(r.sku_key, 3) = 0x583030", sql)
        self.assertIn("WHEN LEFT(r.sku_key, 2) = 0x4230", sql)
        self.assertIn("WHEN LOCATE(0x2D, r.sku_key) > 0 THEN r.sku", sql)
        self.assertIn("SELECT msku, MIN(order_date) AS first_sale_date", sql)
        self.assertIn("GROUP BY msku", sql)
        self.assertIn("MIN(f.first_sale_date) AS opening_date", sql)
        self.assertNotIn("GROUP BY spu", sql)

    def test_field_id_discovery_keeps_empty_custom_field_definitions(self) -> None:
        ids = tagging.field_ids_from_fields(
            [{"id": "123456", "name": "颜色体系", "val_text": ""}]
        )
        self.assertEqual(ids, {"123456"})

    def test_beijing_low_peak_window(self) -> None:
        self.assertTrue(
            tagging.is_beijing_low_peak(datetime(2026, 1, 1, 17, 30, tzinfo=timezone.utc))
        )
        self.assertFalse(
            tagging.is_beijing_low_peak(datetime(2026, 1, 1, 23, 0, tzinfo=timezone.utc))
        )


class ColorSystemWorkbookTest(unittest.TestCase):
    def sample_rows(self) -> list[dict[str, object]]:
        return [
            {
                "SKU": "AA100-BK-S",
                "关联 MSKU": "AA100-BK-S",
                "SPU": "AA100",
                "开售日期": date(2024, 6, 30),
                "首销店铺": "sy_us",
                "拟打值": "A2023",
                "近半年是否有销量": "是",
            },
            {
                "SKU": "BB200-WH-M",
                "关联 MSKU": "BB200-WH-M",
                "SPU": "BB200",
                "开售日期": None,
                "首销店铺": "",
                "拟打值": "待定",
                "近半年是否有销量": "否",
            },
        ]

    def test_dryrun_workbook_has_exact_two_sheets_and_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "dryrun.xlsx"
            tagging.create_dryrun_workbook(self.sample_rows(), output)
            workbook = load_workbook(output, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["拟打标明细", "汇总"])
                detail = workbook["拟打标明细"]
                self.assertEqual([cell.value for cell in detail[1]], tagging.DETAIL_HEADERS)
                summary = workbook["汇总"]
                self.assertEqual(summary["A2"].value, "A2023")
                self.assertEqual(summary["B2"].value, 1)
                self.assertEqual(summary["A3"].value, "待定")
                self.assertEqual(summary["B3"].value, 1)
            finally:
                workbook.close()

    def test_review_file_rejects_invalid_value_and_duplicate_sku(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "review.xlsx"
            tagging.create_dryrun_workbook(self.sample_rows(), output)
            workbook = load_workbook(output)
            sheet = workbook["拟打标明细"]
            sheet["F2"] = "B2024"
            workbook.save(output)
            workbook.close()
            with self.assertRaisesRegex(ValueError, "只允许 A2023/待定"):
                tagging.load_review_file(output)

            tagging.create_dryrun_workbook(self.sample_rows(), output)
            workbook = load_workbook(output)
            sheet = workbook["拟打标明细"]
            sheet["A3"] = "AA100-BK-S"
            workbook.save(output)
            workbook.close()
            with self.assertRaisesRegex(ValueError, "重复 SKU"):
                tagging.load_review_file(output)


class ColorSystemApplyTest(unittest.IsolatedAsyncioTestCase):
    async def test_live_read_can_use_fixed_30_second_retry_policy(self) -> None:
        request = AsyncMock(
            return_value=({"code": 0, "data": [{"sku": "AA100-BK-S"}]}, "token-1")
        )
        with patch.object(product_write_guard, "request_with_retry", new=request):
            products, token = await product_write_guard.fetch_live_products(
                MagicMock(),
                "token-1",
                ["AA100-BK-S"],
                max_retries=3,
                retry_base_seconds=30.0,
                retry_max_seconds=30.0,
            )

        self.assertIn("AA100-BK-S", products)
        self.assertEqual(token, "token-1")
        self.assertEqual(request.await_args.kwargs["max_retries"], 3)
        self.assertEqual(request.await_args.kwargs["retry_base_seconds"], 30.0)
        self.assertEqual(request.await_args.kwargs["retry_max_seconds"], 30.0)

    async def test_apply_uses_guarded_merge_and_exact_rate_limit_policy(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            review = Path(tmp) / "review.xlsx"
            tagging.create_dryrun_workbook(
                [
                    {
                        "SKU": "AA100-BK-S",
                        "关联 MSKU": "AA100-BK-S",
                        "SPU": "AA100",
                        "开售日期": date(2024, 6, 30),
                        "首销店铺": "sy_us",
                        "拟打值": "A2023",
                        "近半年是否有销量": "是",
                    }
                ],
                review,
            )
            args = tagging.parse_args(
                [
                    "--apply",
                    "--review-file",
                    str(review),
                    "--field-id",
                    "999",
                    "--allow-outside-low-peak",
                    "--delay",
                    "0",
                ]
            )
            live_product = {
                "sku": "AA100-BK-S",
                "product_name": "AA100 product",
                "category_id": 7,
                "category": "Top",
                "custom_fields": [
                    {"id": "111", "name": "季节", "val_text": "春夏"},
                    {"id": "999", "name": "颜色体系", "val_text": ""},
                ],
            }
            fake_db = MagicMock()
            fake_client = MagicMock()
            fake_client.generate_token = AsyncMock(return_value=SimpleNamespace(token="token-1"))
            fake_client.aclose = AsyncMock()
            write = AsyncMock(return_value=({"code": 0}, "token-1"))

            with (
                patch.object(tagging, "Database", return_value=fake_db),
                patch.object(tagging, "LingxingClient", return_value=fake_client),
                patch.object(
                    tagging,
                    "fetch_live_products",
                    new=AsyncMock(return_value=({"AA100-BK-S": live_product}, "token-1")),
                ),
                patch.object(tagging, "request_with_retry", new=write),
            ):
                exit_code = await tagging.apply_review_file(args)

            self.assertEqual(exit_code, 0)
            write.assert_awaited_once()
            call = write.await_args
            body = call.args[3]
            self.assertEqual(body["product_name"], "AA100 product")
            self.assertEqual(
                body["custom_fields"],
                [
                    {"id": "111", "name": "季节", "val": "春夏"},
                    {"id": "999", "name": "颜色体系", "val": "A2023"},
                ],
            )
            self.assertEqual(call.kwargs["max_retries"], 3)
            self.assertEqual(call.kwargs["retry_base_seconds"], 30.0)
            self.assertEqual(call.kwargs["retry_max_seconds"], 30.0)

    async def test_apply_records_final_103_and_continues_other_skus(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            review = tmp_path / "review.xlsx"
            rows = []
            for sku in ("AA100-BK-S", "BB200-WH-M"):
                rows.append(
                    {
                        "SKU": sku,
                        "关联 MSKU": sku,
                        "SPU": sku.split("-", 1)[0],
                        "开售日期": date(2024, 6, 30),
                        "首销店铺": "sy_us",
                        "拟打值": "A2023",
                        "近半年是否有销量": "是",
                    }
                )
            tagging.create_dryrun_workbook(rows, review)
            args = tagging.parse_args(
                [
                    "--apply",
                    "--review-file",
                    str(review),
                    "--field-id",
                    "999",
                    "--allow-outside-low-peak",
                    "--delay",
                    "0",
                ]
            )
            live_map = {
                sku: {
                    "sku": sku,
                    "product_name": f"{sku} product",
                    "custom_fields": [
                        {"id": "999", "name": "颜色体系", "val_text": ""}
                    ],
                }
                for sku in ("AA100-BK-S", "BB200-WH-M")
            }
            fake_db = MagicMock()
            fake_client = MagicMock()
            fake_client.generate_token = AsyncMock(return_value=SimpleNamespace(token="token-1"))
            fake_client.aclose = AsyncMock()
            write = AsyncMock(
                side_effect=[
                    (
                        {
                            "code": "103",
                            "message": "请求过于频繁",
                            "request_id": "req-103",
                        },
                        "token-1",
                    ),
                    ({"code": "0"}, "token-1"),
                ]
            )

            with (
                patch.object(tagging, "Database", return_value=fake_db),
                patch.object(tagging, "LingxingClient", return_value=fake_client),
                patch.object(
                    tagging,
                    "fetch_live_products",
                    new=AsyncMock(return_value=(live_map, "token-1")),
                ),
                patch.object(tagging, "request_with_retry", new=write),
                patch.object(tagging, "OUTPUT_DIR", tmp_path / "reports"),
            ):
                exit_code = await tagging.apply_review_file(args)

            self.assertEqual(exit_code, 1)
            self.assertEqual(write.await_count, 2)
            failure_files = list((tmp_path / "reports").glob("*.xlsx"))
            self.assertEqual(len(failure_files), 1)
            workbook = load_workbook(failure_files[0], data_only=True)
            try:
                failure_rows = list(workbook["失败清单"].values)
                self.assertEqual(len(failure_rows), 2)
                self.assertEqual(failure_rows[1][0], "AA100-BK-S")
                self.assertEqual(failure_rows[1][3], "103")
                self.assertEqual(failure_rows[1][5], "req-103")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
