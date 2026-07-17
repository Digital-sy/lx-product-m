from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

from openpyxl import load_workbook

from scripts import color_system_tagging as round1
from scripts import color_system_tagging_round2 as round2


class Round2RulesTest(unittest.TestCase):
    def test_real_value_domain_rules_are_exact(self) -> None:
        self.assertEqual(round2.classify_development_year("历史"), ("A2023", round2.STATUS_CONVERT))
        self.assertEqual(round2.classify_development_year("2024"), ("A2023", round2.STATUS_CONVERT))
        self.assertEqual(
            round2.classify_development_year("2025"),
            ("待定", round2.STATUS_KEEP_FUTURE),
        )
        self.assertEqual(
            round2.classify_development_year("2026"),
            ("待定", round2.STATUS_KEEP_FUTURE),
        )
        self.assertEqual(round2.classify_development_year(""), ("待定", round2.STATUS_KEEP_EMPTY))
        for dirty in ("2024 ", " 2024", "歷史", "2025 ", "2023", "unknown"):
            with self.subTest(dirty=dirty):
                self.assertEqual(
                    round2.classify_development_year(dirty),
                    ("待定", round2.STATUS_KEEP_UNEXPECTED),
                )

    def test_extract_development_year_preserves_dirty_whitespace(self) -> None:
        fields = [
            {
                "id": round2.DEVELOPMENT_YEAR_FIELD_ID,
                "name": "开发年份",
                "val_text": "2024 ",
            }
        ]
        self.assertEqual(round2.extract_development_year(fields), "2024 ")

    def test_extract_development_year_uses_nonempty_fallback_value(self) -> None:
        fields = [
            {
                "id": round2.DEVELOPMENT_YEAR_FIELD_ID,
                "name": "开发年份",
                "val_text": "",
                "val": "2024",
            }
        ]
        self.assertEqual(round2.extract_development_year(fields), "2024")

    def test_extract_development_year_reports_multiple_distinct_values(self) -> None:
        fields = [
            {"id": round2.DEVELOPMENT_YEAR_FIELD_ID, "name": "开发年份", "val": "历史"},
            {"id": round2.DEVELOPMENT_YEAR_FIELD_ID, "name": "开发年份", "val": "2024"},
        ]
        self.assertEqual(round2.extract_development_year(fields), "[多值]历史|2024")

    def test_first_round_file_locks_scope_to_pending_only(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "round1.xlsx"
            round1.create_dryrun_workbook(
                [
                    {
                        "SKU": "AA100-BK-S",
                        "关联 MSKU": "AA100-BK-S",
                        "SPU": "AA100",
                        "开售日期": None,
                        "首销店铺": "",
                        "拟打值": "待定",
                        "近半年是否有销量": "否",
                    },
                    {
                        "SKU": "BB200-WH-M",
                        "关联 MSKU": "BB200-WH-M",
                        "SPU": "BB200",
                        "开售日期": None,
                        "首销店铺": "",
                        "拟打值": "A2023",
                        "近半年是否有销量": "是",
                    },
                ],
                path,
            )
            self.assertEqual(
                round2.load_first_round_pending_skus(path, expected_count=1),
                ["AA100-BK-S"],
            )
            with self.assertRaisesRegex(RuntimeError, "数量不符"):
                round2.load_first_round_pending_skus(path, expected_count=9823)

    def test_group_actual_values_keeps_dirty_variants_separate(self) -> None:
        counts = round2.group_actual_values(
            ["A", "B", "C", "D"],
            {"A": "2024", "B": "2024 ", "C": "2024", "D": ""},
        )
        self.assertEqual(counts, {"2024": 2, "2024 ": 1, "": 1})


class Round2WorkbookTest(unittest.TestCase):
    def test_workbook_contains_summary_distribution_and_unexpected_values(self) -> None:
        rows = round2.prepare_round2_rows(
            ["A", "B", "C", "D", "E"],
            {"A": "历史", "B": "2024", "C": "2025", "D": "", "E": "2024 "},
        )
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "round2.xlsx"
            round2.create_round2_workbook(rows, output)
            workbook = load_workbook(output, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["拟打标明细", "汇总"])
                detail = workbook["拟打标明细"]
                self.assertEqual([cell.value for cell in detail[1]], round2.DETAIL_HEADERS)
                summary_values = list(workbook["汇总"].values)
                self.assertIn(("转 A2023", 2, None), summary_values)
                self.assertIn(("剩余待定", 3, None), summary_values)
                self.assertIn(("空值", 1, None), summary_values)
                self.assertIn(("意外值", 1, None), summary_values)
                self.assertIn(("2024 ", 1, round2.STATUS_KEEP_UNEXPECTED), summary_values)
                self.assertIn(("2024 ", 1, "E"), summary_values)
            finally:
                workbook.close()


class Round2ApplyTest(unittest.IsolatedAsyncioTestCase):
    async def test_second_round_apply_only_writes_a2023_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            review = Path(tmp) / "round2_review.xlsx"
            rows = round2.prepare_round2_rows(
                ["AA100-BK-S", "BB200-WH-M"],
                {"AA100-BK-S": "2024", "BB200-WH-M": "2025"},
            )
            round2.create_round2_workbook(rows, review)
            args = round2.parse_args(
                [
                    "--apply",
                    "--review-file",
                    str(review),
                    "--field-id",
                    "999",
                    "--allow-outside-low-peak",
                    "--delay",
                    "0",
                ]
            )
            live_product = {
                "sku": "AA100-BK-S",
                "product_name": "AA100 product",
                "custom_fields": [
                    {"id": "999", "name": "颜色体系", "val_text": "待定"}
                ],
            }
            fake_db = MagicMock()
            fake_client = MagicMock()
            fake_client.generate_token = AsyncMock(return_value=SimpleNamespace(token="token-1"))
            fake_client.aclose = AsyncMock()
            live_read = AsyncMock(return_value=({"AA100-BK-S": live_product}, "token-1"))
            write = AsyncMock(return_value=({"code": "0"}, "token-1"))

            with (
                patch.object(round1, "Database", return_value=fake_db),
                patch.object(round1, "LingxingClient", return_value=fake_client),
                patch.object(round1, "fetch_live_products", new=live_read),
                patch.object(round1, "request_with_retry", new=write),
            ):
                exit_code = await round1.apply_review_file(
                    args,
                    write_values={"A2023"},
                    failure_prefix="color_system_tagging_round2_failures",
                )

            self.assertEqual(exit_code, 0)
            self.assertEqual(live_read.await_args.args[2], ["AA100-BK-S"])
            write.assert_awaited_once()
            self.assertEqual(write.await_args.args[3]["custom_fields"][-1]["val"], "A2023")


if __name__ == "__main__":
    unittest.main()
