from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts import color_system_high_confidence_all as all_writer


def make_row(**overrides: str) -> dict[str, str]:
    row = {
        "SKU": "SPU1-BK-M",
        "产品名称": "测试黑色上衣",
        "SPU": "SPU1",
        "当前颜色体系": "",
        "SKU结构": "常规",
        "是否多色套装": "否",
        "颜色代码": "BK",
        "品名识别颜色": "黑色",
        "拟打颜色体系": "A2023",
        "置信度": "高",
        "判定类型": "精确唯一匹配",
        "匹配方式": "代码+品名主映射",
        "处理优先级": "P1-有销量",
        "是否允许未来自动写入": "是",
    }
    row.update(overrides)
    return row


class EligibilityTest(unittest.TestCase):
    def test_a2023_b2024_and_p3_are_included(self) -> None:
        self.assertEqual(all_writer.eligibility_reason(make_row()), "")
        self.assertEqual(
            all_writer.eligibility_reason(
                make_row(
                    SKU="SPU2-NB-L",
                    SPU="SPU2",
                    拟打颜色体系="B2024",
                    处理优先级="P3-无销量无Listing",
                    是否允许未来自动写入="否-低优先级",
                )
            ),
            "",
        )

    def test_special_pd_is_included_but_unsafe_rows_are_excluded(self) -> None:
        self.assertEqual(
            all_writer.eligibility_reason(
                make_row(SKU="SPU3-BKPD-M", SKU结构="特殊后缀颜色-PD（绑带款）")
            ),
            "",
        )
        self.assertEqual(
            all_writer.eligibility_reason(make_row(是否多色套装="是")),
            "多色套装",
        )
        self.assertEqual(
            all_writer.eligibility_reason(make_row(SKU结构="未收录颜色代码")),
            "非安全SKU结构",
        )
        self.assertEqual(
            all_writer.eligibility_reason(make_row(品名识别颜色="黑色、白色")),
            "品名识别多个颜色",
        )


class WorkbookTest(unittest.TestCase):
    def test_review_workbook_supports_both_values(self) -> None:
        selected = [
            make_row(),
            make_row(
                SKU="SPU2-NB-L",
                SPU="SPU2",
                颜色代码="NB",
                品名识别颜色="藏青色",
                拟打颜色体系="B2024",
            ),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "review.xlsx"
            source = Path(tmp) / "source.xlsx"
            all_writer.create_review_workbook(
                selected,
                all_writer.Counter(),
                source,
                output,
            )
            wb = load_workbook(output, read_only=True, data_only=True)
            try:
                self.assertIn("拟打标明细", wb.sheetnames)
                ws = wb["拟打标明细"]
                self.assertEqual(ws["B2"].value, "A2023")
                self.assertEqual(ws["B3"].value, "B2024")
            finally:
                wb.close()

            targets = all_writer.load_review_file(output)
            self.assertEqual(
                targets,
                [
                    {"SKU": "SPU1-BK-M", "拟打值": "A2023"},
                    {"SKU": "SPU2-NB-L", "拟打值": "B2024"},
                ],
            )


if __name__ == "__main__":
    unittest.main()
