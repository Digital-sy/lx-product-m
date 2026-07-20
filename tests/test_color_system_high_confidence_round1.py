from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts import color_system_high_confidence_round1 as round1


def make_row(**overrides: str) -> dict[str, str]:
    row = {
        "SKU": "SPU1-BK-M",
        "产品名称": "测试黑色上衣",
        "SPU": "SPU1",
        "当前颜色体系": "",
        "SKU结构": "常规",
        "是否多色套装": "否",
        "颜色代码": "BK",
        "品名识别颜色": "黑色",
        "拟打颜色体系": "A2023",
        "置信度": "高",
        "判定类型": "精确唯一匹配",
        "匹配方式": "代码+品名主映射",
        "处理优先级": "P1-有销量",
        "是否允许未来自动写入": "是",
    }
    row.update(overrides)
    return row


class EligibilityTest(unittest.TestCase):
    def test_safe_row_is_eligible(self) -> None:
        self.assertEqual(round1.eligibility_reason(make_row()), "")

    def test_b2024_p3_special_and_multiple_name_colors_are_excluded(self) -> None:
        self.assertEqual(
            round1.eligibility_reason(make_row(拟打颜色体系="B2024")),
            "不是A2023",
        )
        self.assertEqual(
            round1.eligibility_reason(make_row(处理优先级="P3-无销量无Listing")),
            "P3低优先级",
        )
        self.assertEqual(
            round1.eligibility_reason(
                make_row(
                    处理优先级="P3-无销量无Listing",
                    是否允许未来自动写入="否-低优先级",
                )
            ),
            "分析结果未允许自动写入",
        )
        self.assertEqual(
            round1.eligibility_reason(make_row(SKU结构="特殊后缀颜色-PD（绑带款）")),
            "特殊或异常SKU结构",
        )
        self.assertEqual(
            round1.eligibility_reason(make_row(品名识别颜色="黑色、白色")),
            "品名识别多个颜色",
        )

    def test_selection_prefers_distinct_spu_and_p1(self) -> None:
        rows = [
            make_row(SKU="A-BK-S", SPU="A"),
            make_row(SKU="A-BK-M", SPU="A"),
            make_row(SKU="B-WH-S", SPU="B", 颜色代码="WH", 品名识别颜色="白色"),
            make_row(
                SKU="C-GY-S",
                SPU="C",
                颜色代码="GY",
                品名识别颜色="灰色",
                处理优先级="P2-无销量有Listing",
            ),
        ]
        selected, _, eligible = round1.select_review_rows(rows, limit=3)
        self.assertEqual(eligible, 4)
        self.assertEqual(len(selected), 3)
        self.assertEqual(len({row["SPU"] for row in selected}), 3)
        self.assertTrue(selected[0]["处理优先级"].startswith("P1-"))


class LiveGuardTest(unittest.TestCase):
    def test_live_color_values(self) -> None:
        fields = [
            {"id": "x", "name": "其他", "val": "abc"},
            {
                "id": round1.DEFAULT_FIELD_ID,
                "name": round1.COLOR_FIELD_NAME,
                "val": "B2024",
            },
        ]
        self.assertEqual(
            round1.live_color_values(fields, round1.DEFAULT_FIELD_ID),
            {"B2024"},
        )


class WorkbookTest(unittest.TestCase):
    def test_review_workbook_can_be_loaded_by_existing_writer(self) -> None:
        selected = [make_row()]
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "review.xlsx"
            source = Path(tmp) / "source.xlsx"
            round1.create_review_workbook(selected, round1.Counter(), 1, source, output)
            wb = load_workbook(output, read_only=True, data_only=True)
            try:
                self.assertIn("拟打标明细", wb.sheetnames)
                ws = wb["拟打标明细"]
                self.assertEqual(ws["A2"].value, "SPU1-BK-M")
                self.assertEqual(ws["B2"].value, "A2023")
            finally:
                wb.close()

            targets = round1.load_review_file(output)
            self.assertEqual(targets, [{"SKU": "SPU1-BK-M", "拟打值": "A2023"}])


if __name__ == "__main__":
    unittest.main()
